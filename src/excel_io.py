from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


JOB_TABLE = "JobInputV5Table"
SERVICE_TABLE = "ServiceInputV5Table"
JOB_REQUIRED_HEADERS = {"job_ref", "bot_status", "bot_result"}
SERVICE_REQUIRED_HEADERS = {
    "job_ref",
    "service_seq",
    "service_code",
    "service_status",
    "service_result",
}

COMMITTING_TVC = "COMMITTING_TVC"
TVC_SAVED_PENDING_EXCEL = "TVC_SAVED_PENDING_EXCEL"
UNCERTAIN_TVC_SAVE = "UNCERTAIN_TVC_SAVE: กรุณาตรวจสอบใน T.V.C ก่อน retry"
DIRTY_TVC_FORM_POSSIBLE = (
    "DIRTY_TVC_FORM_POSSIBLE: กรุณาตรวจสอบ/ปิดหน้าใบงาน T.V.C ก่อนเริ่มใหม่"
)
COMMIT_MARKERS = {COMMITTING_TVC, TVC_SAVED_PENDING_EXCEL}


def _table_rows(ws, table_name):
    if table_name not in ws.tables:
        raise RuntimeError(f"ไม่พบ Excel Table {table_name}")
    min_col,min_row,max_col,max_row = range_boundaries(ws.tables[table_name].ref)
    headers=[ws.cell(min_row,c).value for c in range(min_col,max_col+1)]
    for row_idx in range(min_row+1,max_row+1):
        vals=[ws.cell(row_idx,c).value for c in range(min_col,max_col+1)]
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        d={headers[i]:vals[i] for i in range(len(headers))}
        d["_row"]=row_idx
        yield d


def _table_header_columns(ws,table_name):
    if table_name not in ws.tables:
        raise RuntimeError(f"ไม่พบ Excel Table {table_name}")
    min_col,min_row,max_col,_=range_boundaries(ws.tables[table_name].ref)
    return {
        str(ws.cell(min_row,c).value or "").strip():c
        for c in range(min_col,max_col+1)
    }


def load_jobs(path: Path, job_sheet="JOB_INPUT", service_sheet="SERVICE_INPUT"):
    wb=load_workbook(path)
    try:
        jws=wb[job_sheet]; sws=wb[service_sheet]
        services=list(_table_rows(sws,SERVICE_TABLE))
        jobs=[]
        for job in _table_rows(jws,JOB_TABLE):
            ref=str(job.get("job_ref") or "").strip()
            status=str(job.get("bot_status") or "").strip().upper()
            if not ref or status!="WAIT":
                continue
            matched=[s for s in services
                     if str(s.get("job_ref") or "").strip()==ref
                     and str(s.get("service_status") or "").strip().upper() in {"WAIT",""}]
            matched.sort(key=lambda x:int(x.get("service_seq") or 999999))
            job["_services"]=matched
            jobs.append(job)
        return jobs
    finally:
        wb.close()


def validate_workbook(path: Path, job_sheet="JOB_INPUT", service_sheet="SERVICE_INPUT"):
    """Validate the existing workbook contract without changing the file."""
    path=Path(path)
    if path.suffix.lower()!=".xlsx":
        raise ValueError("รองรับเฉพาะไฟล์ Excel .xlsx")
    if not path.is_file():
        raise FileNotFoundError(f"ไม่พบไฟล์ Excel: {path}")

    wb=load_workbook(path,read_only=False,data_only=False)
    try:
        missing_sheets=[name for name in (job_sheet,service_sheet) if name not in wb.sheetnames]
        if missing_sheets:
            raise RuntimeError(f"ไม่พบ Sheet: {', '.join(missing_sheets)}")

        jws=wb[job_sheet]; sws=wb[service_sheet]
        for ws,table_name,required in (
            (jws,JOB_TABLE,JOB_REQUIRED_HEADERS),
            (sws,SERVICE_TABLE,SERVICE_REQUIRED_HEADERS),
        ):
            if table_name not in ws.tables:
                raise RuntimeError(f"ไม่พบ Excel Table {table_name}")
            headers=set(_table_header_columns(ws,table_name))
            missing=sorted(required-headers)
            if missing:
                raise RuntimeError(f"Table {table_name} ขาดคอลัมน์: {', '.join(missing)}")
    finally:
        wb.close()
    return True


def get_job_stats(path: Path,job_sheet="JOB_INPUT"):
    """Return read-only JOB counts used by the desktop controller."""
    wb=load_workbook(path,read_only=False,data_only=True)
    try:
        if job_sheet not in wb.sheetnames:
            raise RuntimeError(f"ไม่พบ Sheet: {job_sheet}")
        ws=wb[job_sheet]
        counts={"WAIT":0,"DONE":0,"ERROR":0,"RUNNING":0,"OTHER":0}
        total=0
        for job in _table_rows(ws,JOB_TABLE):
            ref=str(job.get("job_ref") or "").strip()
            if not ref:
                continue
            total+=1
            status=str(job.get("bot_status") or "").strip().upper()
            if status in counts and status!="OTHER":
                counts[status]+=1
            else:
                counts["OTHER"]+=1
        counts["TOTAL"]=total
        counts["COMPLETED"]=counts["DONE"]+counts["ERROR"]
        return counts
    finally:
        wb.close()


def get_job_errors(path: Path,job_sheet="JOB_INPUT"):
    """Return read-only ERROR job details for the GUI queue summary."""
    wb=load_workbook(path,read_only=False,data_only=True)
    try:
        if job_sheet not in wb.sheetnames:
            raise RuntimeError(f"ไม่พบ Sheet: {job_sheet}")
        ws=wb[job_sheet]
        errors=[]
        for job in _table_rows(ws,JOB_TABLE):
            if str(job.get("bot_status") or "").strip().upper()!="ERROR":
                continue
            errors.append({
                "job_ref":str(job.get("job_ref") or "").strip(),
                "bot_result":str(job.get("bot_result") or "").strip(),
            })
        return errors
    finally:
        wb.close()


def get_safety_issues(path: Path,job_sheet="JOB_INPUT"):
    """Return workbook-derived locks without changing the workbook."""
    wb=load_workbook(path,read_only=False,data_only=True)
    try:
        if job_sheet not in wb.sheetnames:
            raise RuntimeError(f"ไม่พบ Sheet: {job_sheet}")
        ws=wb[job_sheet]
        issues=[]
        for job in _table_rows(ws,JOB_TABLE):
            ref=str(job.get("job_ref") or "").strip()
            status=str(job.get("bot_status") or "").strip().upper()
            result=str(job.get("bot_result") or "").strip()
            if status=="ERROR" and "UNCERTAIN_TVC_SAVE" in result.upper():
                issues.append({
                    "reason":"UNCERTAIN_TVC_SAVE",
                    "job_ref":ref,
                    "bot_status":status,
                    "bot_result":result,
                    "message":(
                        "พบงานที่ไม่ยืนยันผลการบันทึก T.V.C "
                        "กรุณาตรวจสอบก่อนเริ่มใหม่"
                    ),
                })
        return issues
    finally:
        wb.close()


def _read_recovery_snapshot(path,job_sheet,service_sheet):
    """Read recovery state without saving or mutating the workbook."""
    wb=load_workbook(path,read_only=False,data_only=False)
    try:
        jws=wb[job_sheet]; sws=wb[service_sheet]
        return {
            "jobs":list(_table_rows(jws,JOB_TABLE)),
            "services":list(_table_rows(sws,SERVICE_TABLE)),
        }
    finally:
        wb.close()


def _base_recovery_result(outcome,message=""):
    return {
        "outcome":outcome,
        "message":message,
        "job_ref":"",
        "previous_status":"",
        "previous_result":"",
        "job_reset":False,
        "services_reset":0,
        "verified":False,
        "running_count":0,
        "reference_stale":False,
        "warning":"",
    }


def _select_recovery_target(jobs,current_job_ref="",expected_no_active=False):
    running=[
        job for job in jobs
        if str(job.get("bot_status") or "").strip().upper()=="RUNNING"
    ]
    # RUNNING rows are the source of truth. current_job_ref is only a hint
    # because the GUI may not have received the next job_start event yet.
    if len(running)>1:
        return "ambiguous",None,running
    if len(running)==1:
        return "target",running[0],running
    return "already_clean",None,running


def inspect_recovery_state(
    path: Path,
    current_job_ref="",
    job_sheet="JOB_INPUT",
    service_sheet="SERVICE_INPUT",
    expected_no_active=False,
):
    """Inspect the workbook first; this function never saves it."""
    snapshot=_read_recovery_snapshot(path,job_sheet,service_sheet)
    selection,target,running=_select_recovery_target(
        snapshot["jobs"],
        current_job_ref,
        expected_no_active,
    )
    result=_base_recovery_result(selection)
    result["running_count"]=len(running)

    if target is None:
        result["verified"]=selection=="already_clean"
        if selection=="ambiguous":
            result["message"]="พบ JOB RUNNING หลายรายการ ต้องตรวจ Excel ก่อนเริ่มใหม่"
        else:
            result["message"]="ไม่พบ JOB ที่ค้างเป็น RUNNING"
            requested_ref=str(current_job_ref or "").strip()
            if requested_ref:
                matches=[
                    job for job in snapshot["jobs"]
                    if str(job.get("job_ref") or "").strip()==requested_ref
                ]
                if len(matches)==1:
                    marker=str(matches[0].get("bot_result") or "").strip()
                    if marker in COMMIT_MARKERS or marker.startswith("UNCERTAIN_TVC_SAVE:"):
                        result.update({
                            "outcome":"uncertain_commit",
                            "job_ref":requested_ref,
                            "previous_status":str(matches[0].get("bot_status") or "").strip().upper(),
                            "previous_result":marker,
                            "message":UNCERTAIN_TVC_SAVE,
                        })
        return result

    ref=str(target.get("job_ref") or "").strip()
    status=str(target.get("bot_status") or "").strip().upper()
    marker=str(target.get("bot_result") or "").strip()
    requested_ref=str(current_job_ref or "").strip()
    reference_stale=bool(requested_ref and requested_ref!=ref)
    result.update({
        "job_ref":ref,
        "previous_status":status,
        "previous_result":marker,
        "reference_stale":reference_stale,
        "warning":(
            f"current_job_ref={requested_ref} ไม่ตรงกับ JOB RUNNING={ref}; ใช้ JOB RUNNING เป็น target"
            if reference_stale else ""
        ),
    })

    if status=="RUNNING":
        result["outcome"]="pending_recovery"
        result["message"]="พบ JOB ค้างเป็น RUNNING"
        return result

    result["verified"]=True
    if marker in COMMIT_MARKERS or marker.startswith("UNCERTAIN_TVC_SAVE:"):
        result["outcome"]="uncertain_commit"
        result["message"]=UNCERTAIN_TVC_SAVE
    else:
        result["outcome"]="already_clean"
        result["message"]=f"JOB อยู่ในสถานะ {status or 'ว่าง'}"
    return result


def reconcile_process_exit(
    path: Path,
    current_job_ref="",
    job_sheet="JOB_INPUT",
    service_sheet="SERVICE_INPUT",
    precommit_result="process จบผิดปกติก่อน save",
    expected_no_active=False,
):
    """Reconcile and then verify Excel after any abnormal subprocess exit."""
    inspection=inspect_recovery_state(
        path,
        current_job_ref,
        job_sheet,
        service_sheet,
        expected_no_active,
    )
    if inspection["outcome"]!="pending_recovery":
        return inspection

    ref=inspection["job_ref"]
    wb=load_workbook(path)
    outcome="recovered"
    job_reset=False
    services_reset=0
    previous_status=""
    previous_result=""
    try:
        jws=wb[job_sheet]; sws=wb[service_sheet]
        running_jobs=[
            job for job in _table_rows(jws,JOB_TABLE)
            if str(job.get("bot_status") or "").strip().upper()=="RUNNING"
        ]
        if len(running_jobs)!=1:
            return inspect_recovery_state(
                path,
                current_job_ref,
                job_sheet,
                service_sheet,
                expected_no_active,
            )

        target=running_jobs[0]
        ref=str(target.get("job_ref") or "").strip()
        previous_status=str(target.get("bot_status") or "").strip().upper()
        previous_result=str(target.get("bot_result") or "").strip()

        job_columns=_table_header_columns(jws,JOB_TABLE)
        row=int(target["_row"])
        uncertain=(
            previous_result in COMMIT_MARKERS
            or previous_result.startswith("UNCERTAIN_TVC_SAVE:")
        )
        if uncertain:
            outcome="uncertain_commit"
            jws.cell(row=row,column=job_columns["bot_status"]).value="ERROR"
            jws.cell(row=row,column=job_columns["bot_result"]).value=UNCERTAIN_TVC_SAVE
            job_reset=True
        else:
            outcome="recovered"
            jws.cell(row=row,column=job_columns["bot_status"]).value="WAIT"
            jws.cell(row=row,column=job_columns["bot_result"]).value=precommit_result
            job_reset=True
            service_columns=_table_header_columns(sws,SERVICE_TABLE)
            for service in _table_rows(sws,SERVICE_TABLE):
                if str(service.get("job_ref") or "").strip()!=ref:
                    continue
                if str(service.get("service_status") or "").strip().upper()!="ADDED":
                    continue
                service_row=int(service["_row"])
                sws.cell(row=service_row,column=service_columns["service_status"]).value="WAIT"
                sws.cell(row=service_row,column=service_columns["service_result"]).value=precommit_result
                services_reset+=1

        wb.save(path)
    finally:
        wb.close()

    verification=inspect_recovery_state(
        path,
        ref,
        job_sheet,
        service_sheet,
        expected_no_active,
    )
    verified=(
        bool(verification["verified"])
        and verification.get("running_count")==0
        and verification.get("outcome") in {"already_clean","uncertain_commit"}
    )
    if outcome=="recovered":
        snapshot=_read_recovery_snapshot(path,job_sheet,service_sheet)
        verified=verified and not any(
            str(service.get("job_ref") or "").strip()==ref
            and str(service.get("service_status") or "").strip().upper()=="ADDED"
            for service in snapshot["services"]
        )

    result=_base_recovery_result(outcome)
    result.update({
        "message":UNCERTAIN_TVC_SAVE if outcome=="uncertain_commit" else precommit_result,
        "job_ref":ref,
        "previous_status":previous_status,
        "previous_result":previous_result,
        "job_reset":job_reset,
        "services_reset":services_reset,
        "verified":verified,
        "running_count":verification.get("running_count",0),
        "reference_stale":inspection.get("reference_stale",False),
        "warning":inspection.get("warning",""),
    })
    if not verified:
        result["outcome"]="failed"
        result["message"]="recovery เสร็จแต่ตรวจพบว่า JOB ยังไม่อยู่ในสถานะปลอดภัย"
    return result


def recover_forced_stop(
    path: Path,
    current_job_ref="",
    job_sheet="JOB_INPUT",
    service_sheet="SERVICE_INPUT",
    result_message="ถูกบังคับหยุดก่อน save",
):
    """Backward-compatible wrapper for the forced-stop recovery entry point."""
    return reconcile_process_exit(
        path,
        current_job_ref,
        job_sheet,
        service_sheet,
        precommit_result=result_message,
    )


def write_job_result(path: Path,row,status,result="",tvc_job_no=""):
    wb=load_workbook(path)
    try:
        ws=wb["JOB_INPUT"]
        h={c.value:i+1 for i,c in enumerate(ws[1])}
        ws.cell(row=row,column=h["bot_status"]).value=status
        ws.cell(row=row,column=h["bot_result"]).value=result
        if tvc_job_no and "tvc_job_no" in h:
            ws.cell(row=row,column=h["tvc_job_no"]).value=tvc_job_no
        wb.save(path)
    finally:
        wb.close()


def write_service_result(path: Path,row,status,result=""):
    wb=load_workbook(path)
    try:
        ws=wb["SERVICE_INPUT"]
        h={c.value:i+1 for i,c in enumerate(ws[1])}
        ws.cell(row=row,column=h["service_status"]).value=status
        ws.cell(row=row,column=h["service_result"]).value=result
        wb.save(path)
    finally:
        wb.close()
