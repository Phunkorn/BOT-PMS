import queue
from pathlib import Path
import subprocess
import sys
import tempfile
import threading
import time
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

SRC=Path(__file__).resolve().parents[1]/"src"
sys.path.insert(0,str(SRC))

import bot
import tvc_control
from excel_io import (
    COMMITTING_TVC,
    TVC_SAVED_PENDING_EXCEL,
    UNCERTAIN_TVC_SAVE,
    inspect_recovery_state,
    reconcile_process_exit,
)

tvc_control.initialize_gui_dependencies()


JOB_HEADERS=["job_ref","bot_status","bot_result"]
SERVICE_HEADERS=[
    "job_ref",
    "service_seq",
    "service_code",
    "service_status",
    "service_result",
]


def create_workbook(path,jobs,services):
    wb=Workbook()
    jws=wb.active
    jws.title="JOB_INPUT"
    jws.append(JOB_HEADERS)
    for row in jobs:
        jws.append([row.get(header,"") for header in JOB_HEADERS])
    job_ref=f"A1:C{max(len(jobs)+1,2)}"
    if not jobs:
        jws.append(["","",""])
    job_table=Table(displayName="JobInputV5Table",ref=job_ref)
    job_table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
    jws.add_table(job_table)

    sws=wb.create_sheet("SERVICE_INPUT")
    sws.append(SERVICE_HEADERS)
    for row in services:
        sws.append([row.get(header,"") for header in SERVICE_HEADERS])
    service_ref=f"A1:E{max(len(services)+1,2)}"
    if not services:
        sws.append(["","","","",""])
    service_table=Table(displayName="ServiceInputV5Table",ref=service_ref)
    service_table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True)
    sws.add_table(service_table)
    wb.save(path)
    wb.close()


def read_rows(path):
    wb=load_workbook(path,data_only=False)
    try:
        jws=wb["JOB_INPUT"]; sws=wb["SERVICE_INPUT"]
        jobs=[
            {JOB_HEADERS[i]:jws.cell(row, i+1).value for i in range(len(JOB_HEADERS))}
            for row in range(2,jws.max_row+1)
            if jws.cell(row,1).value
        ]
        services=[
            {SERVICE_HEADERS[i]:sws.cell(row, i+1).value for i in range(len(SERVICE_HEADERS))}
            for row in range(2,sws.max_row+1)
            if sws.cell(row,1).value
        ]
        return jobs,services
    finally:
        wb.close()


class RecoveryStateTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(prefix="tvc_recovery_tests_")
        self.path=Path(self.tmp.name)/"fixture.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_a_precommit_force_resets_job_and_services(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"RUNNING","bot_result":"Bot กำลังทำงาน"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
        )
        result=reconcile_process_exit(
            self.path,
            "A",
            precommit_result="ถูกบังคับหยุดก่อน save",
        )
        self.assertEqual(result["outcome"],"recovered")
        self.assertTrue(result["verified"])
        jobs,services=read_rows(self.path)
        self.assertEqual(jobs[0]["bot_status"],"WAIT")
        self.assertEqual(jobs[0]["bot_result"],"ถูกบังคับหยุดก่อน save")
        self.assertEqual(services[0]["service_status"],"WAIT")
        self.assertEqual(services[0]["service_result"],"ถูกบังคับหยุดก่อน save")
        verify=inspect_recovery_state(self.path,"A")
        self.assertEqual(verify["outcome"],"already_clean")
        self.assertTrue(verify["verified"])

    def test_b_and_c_commit_markers_become_uncertain_without_service_reset(self):
        for marker in (COMMITTING_TVC,TVC_SAVED_PENDING_EXCEL):
            with self.subTest(marker=marker):
                create_workbook(
                    self.path,
                    [{"job_ref":"A","bot_status":"RUNNING","bot_result":marker}],
                    [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
                )
                result=reconcile_process_exit(self.path,"A")
                self.assertEqual(result["outcome"],"uncertain_commit")
                self.assertTrue(result["verified"])
                self.assertEqual(result["services_reset"],0)
                jobs,services=read_rows(self.path)
                self.assertEqual(jobs[0]["bot_status"],"ERROR")
                self.assertEqual(jobs[0]["bot_result"],UNCERTAIN_TVC_SAVE)
                self.assertEqual(services[0]["service_status"],"ADDED")

    def test_f_multiple_running_is_ambiguous_and_unchanged(self):
        create_workbook(
            self.path,
            [
                {"job_ref":"A","bot_status":"RUNNING","bot_result":"working"},
                {"job_ref":"B","bot_status":"RUNNING","bot_result":"working"},
            ],
            [
                {"job_ref":"A","service_seq":1,"service_code":"S1","service_status":"ADDED","service_result":"added"},
                {"job_ref":"B","service_seq":1,"service_code":"S2","service_status":"ADDED","service_result":"added"},
            ],
        )
        result=reconcile_process_exit(self.path)
        self.assertEqual(result["outcome"],"ambiguous")
        self.assertFalse(result["verified"])
        jobs,services=read_rows(self.path)
        self.assertEqual([row["bot_status"] for row in jobs],["RUNNING","RUNNING"])
        self.assertEqual([row["service_status"] for row in services],["ADDED","ADDED"])

    def test_zero_running_is_already_clean_even_when_hint_is_missing(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"WAIT","bot_result":""}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"WAIT","service_result":""}],
        )
        result=reconcile_process_exit(self.path,"MISSING")
        self.assertEqual(result["outcome"],"already_clean")
        self.assertTrue(result["verified"])
        self.assertEqual(result["running_count"],0)

    def test_stale_done_hint_reconciles_the_single_running_job(self):
        create_workbook(
            self.path,
            [
                {"job_ref":"JOB-A","bot_status":"DONE","bot_result":"done"},
                {"job_ref":"JOB-B","bot_status":"RUNNING","bot_result":"Bot กำลังทำงาน"},
            ],
            [
                {"job_ref":"JOB-A","service_seq":1,"service_code":"S1","service_status":"ADDED","service_result":"done"},
                {"job_ref":"JOB-B","service_seq":1,"service_code":"S2","service_status":"ADDED","service_result":"added"},
            ],
        )
        result=reconcile_process_exit(self.path,"JOB-A")
        self.assertEqual(result["outcome"],"recovered")
        self.assertTrue(result["verified"])
        self.assertEqual(result["job_ref"],"JOB-B")
        self.assertTrue(result["reference_stale"])
        self.assertIn("JOB-A",result["warning"])
        self.assertIn("JOB-B",result["warning"])
        jobs,services=read_rows(self.path)
        self.assertEqual([job["bot_status"] for job in jobs],["DONE","WAIT"])
        self.assertEqual([service["service_status"] for service in services],["ADDED","WAIT"])

    def test_no_hint_reconciles_the_single_running_job(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"RUNNING","bot_result":"working"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
        )
        result=reconcile_process_exit(self.path)
        self.assertEqual(result["outcome"],"recovered")
        self.assertTrue(result["verified"])
        self.assertEqual(result["job_ref"],"A")
        self.assertFalse(result["reference_stale"])

    def test_verification_with_running_remaining_is_failed(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"RUNNING","bot_result":"working"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
        )
        initial=inspect_recovery_state(self.path,"A")
        unsafe_verification={
            "outcome":"pending_recovery",
            "message":"ยังมี RUNNING",
            "job_ref":"A",
            "verified":False,
            "running_count":1,
        }
        with mock.patch.object(
            sys.modules["excel_io"],
            "inspect_recovery_state",
            side_effect=[initial,unsafe_verification],
        ):
            result=reconcile_process_exit(self.path,"A")
        self.assertEqual(result["outcome"],"failed")
        self.assertFalse(result["verified"])
        self.assertEqual(result["running_count"],1)


class RuntimeValidationTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(prefix="tvc_runtime_tests_")
        self.root=Path(self.tmp.name)
        self.python_exe=self.root/"python.exe"

    def tearDown(self):
        self.tmp.cleanup()

    @staticmethod
    def _result(returncode=0,stdout="",stderr=""):
        return subprocess.CompletedProcess([],returncode,stdout,stderr)

    def test_python_exe_missing(self):
        with self.assertRaisesRegex(RuntimeError,"ไม่พบ Python"):
            tvc_control.validate_bot_runtime(self.python_exe)

    def test_python_exe_zero_bytes(self):
        self.python_exe.write_bytes(b"")
        with self.assertRaisesRegex(RuntimeError,"ว่างเปล่า"):
            tvc_control.validate_bot_runtime(self.python_exe)

    def test_pythonw_is_never_accepted_for_bot_subprocess(self):
        pythonw=self.root/"pythonw.exe"
        pythonw.write_bytes(b"mock")
        with self.assertRaisesRegex(RuntimeError,"ห้ามใช้ pythonw"):
            tvc_control.validate_bot_runtime(pythonw)

    def test_python_version_command_failure(self):
        self.python_exe.write_bytes(b"mock")
        with mock.patch.object(
            tvc_control.subprocess,
            "run",
            return_value=self._result(1,stderr="broken runtime"),
        ):
            with self.assertRaisesRegex(RuntimeError,"ใช้งานไม่ได้"):
                tvc_control.validate_bot_runtime(self.python_exe)

    def test_dependency_import_failure(self):
        self.python_exe.write_bytes(b"mock")
        with mock.patch.object(
            tvc_control.subprocess,
            "run",
            side_effect=[
                self._result(0,stdout="Python 3.12"),
                self._result(1,stderr="No module named 'pywinauto'"),
            ],
        ):
            with self.assertRaisesRegex(RuntimeError,"ขาด dependencies"):
                tvc_control.validate_bot_runtime(self.python_exe)

    def test_python_version_timeout_has_clear_message(self):
        self.python_exe.write_bytes(b"mock")
        with mock.patch.object(
            tvc_control.subprocess,
            "run",
            side_effect=subprocess.TimeoutExpired("python --version",8),
        ):
            with self.assertRaisesRegex(RuntimeError,"Runtime check timeout.*--version"):
                tvc_control.validate_bot_runtime(self.python_exe)

    def test_dependency_import_timeout_has_clear_message(self):
        self.python_exe.write_bytes(b"mock")
        with mock.patch.object(
            tvc_control.subprocess,
            "run",
            side_effect=[
                self._result(0,stdout="Python 3.12"),
                subprocess.TimeoutExpired("dependency import",8),
            ],
        ):
            with self.assertRaisesRegex(RuntimeError,"Runtime check timeout.*dependencies"):
                tvc_control.validate_bot_runtime(self.python_exe)

    def test_valid_console_runtime_checks_version_and_dependencies(self):
        self.python_exe.write_bytes(b"mock")
        with mock.patch.object(
            tvc_control.subprocess,
            "run",
            side_effect=[
                self._result(0,stdout="Python 3.12"),
                self._result(0),
            ],
        ) as run:
            selected=tvc_control.validate_bot_runtime(self.python_exe)
        self.assertEqual(selected,self.python_exe.resolve())
        self.assertEqual(run.call_args_list[0].args[0],[str(self.python_exe),"--version"])
        dependency_command=run.call_args_list[1].args[0]
        self.assertEqual(dependency_command[:2],[str(self.python_exe),"-c"])
        for module in ("openpyxl","pywinauto","psutil"):
            self.assertIn(module,dependency_command[2])


class FakeVar:
    def __init__(self,value=""):
        self.value=value
    def get(self):
        return self.value
    def set(self,value):
        self.value=value


class FakeButton:
    def __init__(self):
        self.state="disabled"
    def configure(self,**values):
        if "state" in values:
            self.state=values["state"]
    def __getitem__(self,key):
        if key=="state":
            return self.state
        raise KeyError(key)


class FakeRoot:
    def after(self,_delay,callback):
        callback()


class FakeProcess:
    def __init__(self,returncode=None,kill_error=None):
        self.returncode=returncode
        self.kill_error=kill_error
        self.kill_calls=0
        self.pid=12345
    def poll(self):
        return self.returncode
    def kill(self):
        self.kill_calls+=1
        if self.kill_error:
            raise self.kill_error
        self.returncode=-9
    def wait(self,timeout=None):
        if self.returncode is None:
            raise subprocess.TimeoutExpired("fake",timeout)
        return self.returncode


class ImmediateThread:
    def __init__(self,target,args=(),daemon=None):
        self.target=target
        self.args=args
    def start(self):
        self.target(*self.args)


def make_fake_app(path):
    app=object.__new__(tvc_control.TVCControlApp)
    app.process=None
    app.stop_file=None
    app.events=queue.Queue()
    app.valid_excel=True
    app.stop_request_sent=False
    app.stop_event_seen=False
    app.stop_event_phase=""
    app.last_batch_success=None
    app.current_job_ref="A"
    app.force_stop_used=False
    app.force_retry_available=False
    app.recovery_in_progress=False
    app.recovery_failed=False
    app.safety_locks={}
    app.safety_metadata_health=tvc_control.SAFETY_METADATA_MISSING
    app.safety_metadata_error=""
    app.last_return_code=None
    app.closing=False
    app.runtime_check_in_progress=False
    app.runtime_valid=True
    app.runtime_error=""
    app.bot_python=Path(sys.base_prefix)/"python.exe"
    app.excel_var=FakeVar(str(path))
    app.status_var=FakeVar("กำลังทำงาน")
    app.safety_status_var=FakeVar("ไม่ล็อก")
    app.safety_detail_var=FakeVar("")
    app.current_job_var=FakeVar("- / 0")
    app.wait_var=FakeVar("0")
    app.done_var=FakeVar("0")
    app.error_var=FakeVar("0")
    app.progress_text_var=FakeVar("0 / 0")
    app.progress=mock.Mock()
    app.start_button=FakeButton()
    app.stop_button=FakeButton()
    app.browse_button=FakeButton()
    app.runtime_button=FakeButton()
    app.root=FakeRoot()
    app.logs=[]
    app.append_log=app.logs.append
    app._request_stats=lambda:None
    app._set_status=lambda value:app.status_var.set(value)
    return app


class GuiStateTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(prefix="tvc_gui_state_tests_")
        self.path=Path(self.tmp.name)/"fixture.xlsx"

    def tearDown(self):
        self.tmp.cleanup()

    def test_d_nonzero_exit_always_reconciles(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"RUNNING","bot_result":"Bot กำลังทำงาน"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
        )
        app=make_fake_app(self.path)
        process=FakeProcess(returncode=1)
        app.process=process
        with mock.patch.object(tvc_control.threading,"Thread",ImmediateThread):
            app._on_process_end(process,1)
        event=app.events.get_nowait()
        self.assertEqual(event[0],"reconciliation")
        app._on_reconciliation(event[1],event[2])
        jobs,services=read_rows(self.path)
        self.assertEqual(jobs[0]["bot_status"],"WAIT")
        self.assertEqual(services[0]["service_status"],"WAIT")
        self.assertIn("DIRTY_TVC_FORM_POSSIBLE",jobs[0]["bot_result"])
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")
        self.assertTrue(app.recovery_failed)
        self.assertEqual(app.start_button["state"],"disabled")

    def test_d2_force_termination_locks_even_when_excel_is_already_clean(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"WAIT","bot_result":""}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"WAIT","service_result":""}],
        )
        app=make_fake_app(self.path)
        app.force_stop_used=True
        app.safety_locks={}
        app._set_safety_lock=mock.Mock()

        app._on_reconciliation(
            {
                "outcome":"already_clean",
                "message":"ไม่พบ JOB ที่ค้างเป็น RUNNING",
                "verified":True,
                "running_count":0,
                "job_ref":"A",
                "job_reset":False,
                "services_reset":0,
            },
            -9,
        )

        app._set_safety_lock.assert_called_once()
        self.assertEqual(
            app._set_safety_lock.call_args.args[1],
            "DIRTY_TVC_FORM_POSSIBLE",
        )
        self.assertTrue(app.recovery_failed)
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")

    def test_d3_dirty_lock_persist_failure_is_fail_closed_and_not_reported_safe(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"WAIT","bot_result":"dirty"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"WAIT","service_result":"dirty"}],
        )
        app=make_fake_app(self.path)
        app.force_stop_used=True
        app.safety_state_file=Path(self.tmp.name)/"safety_locks.json"
        payload={
            "outcome":"already_clean",
            "message":"ไม่พบ JOB ที่ค้างเป็น RUNNING",
            "verified":True,
            "running_count":0,
            "job_ref":"A",
            "job_reset":False,
            "services_reset":0,
        }
        with mock.patch.object(
            tvc_control,"save_persisted_safety_locks",side_effect=PermissionError("denied")
        ):
            app._on_reconciliation(payload,-9)
        self.assertTrue(app.safety_locks)
        self.assertEqual(
            app.safety_metadata_health,tvc_control.SAFETY_METADATA_WRITE_FAILED
        )
        self.assertEqual(app.status_var.get(),tvc_control.SAFETY_PERSISTENCE_ERROR)
        self.assertTrue(any("ห้าม restart" in line for line in app.logs))
        self.assertTrue(app.recovery_failed)

    def test_e_terminate_failure_uses_kill_then_reenables_stop(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"RUNNING","bot_result":"working"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
        )
        app=make_fake_app(self.path)
        process=FakeProcess(returncode=None,kill_error=OSError("kill failed"))
        app.process=process
        app.stop_request_sent=True
        app.force_stop_used=True
        with mock.patch.object(tvc_control.psutil,"Process",side_effect=OSError("psutil failed")):
            app._terminate_process_tree(process)
        event=app.events.get_nowait()
        self.assertEqual(event[0],"force_termination_failed")
        app._on_force_termination_failed(event[1],event[2])
        self.assertEqual(process.kill_calls,1)
        self.assertEqual(app.status_var.get(),"หยุดไม่สำเร็จ - process ยังทำงาน")
        self.assertTrue(app.force_retry_available)
        self.assertFalse(app.stop_request_sent)
        self.assertEqual(app.stop_button["state"],"normal")
        self.assertEqual(app.start_button["state"],"disabled")

    def test_reload_only_unblocks_after_running_is_fixed(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"RUNNING","bot_result":"working"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
        )
        app=make_fake_app(self.path)
        app.process=None
        self.assertFalse(app._select_excel(self.path,False))
        self.assertTrue(app.recovery_failed)
        self.assertEqual(app.start_button["state"],"disabled")

        wb=load_workbook(self.path)
        wb["JOB_INPUT"].cell(2,2).value="WAIT"
        wb["SERVICE_INPUT"].cell(2,4).value="WAIT"
        wb.save(self.path)
        wb.close()

        self.assertTrue(app._select_excel(self.path,False))
        self.assertFalse(app.recovery_failed)
        self.assertEqual(app.start_button["state"],"normal")

    def test_f_ambiguous_outcome_blocks_start(self):
        create_workbook(
            self.path,
            [
                {"job_ref":"A","bot_status":"RUNNING","bot_result":"working"},
                {"job_ref":"B","bot_status":"RUNNING","bot_result":"working"},
            ],
            [
                {"job_ref":"A","service_seq":1,"service_code":"S1","service_status":"ADDED","service_result":"added"},
                {"job_ref":"B","service_seq":1,"service_code":"S2","service_status":"ADDED","service_result":"added"},
            ],
        )
        result=reconcile_process_exit(self.path)
        app=make_fake_app(self.path)
        app.process=None
        app.current_job_ref=""
        with mock.patch.object(tvc_control.messagebox,"showerror"):
            app._on_reconciliation(result,1)
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")
        self.assertTrue(app.recovery_failed)
        self.assertEqual(app.start_button["state"],"disabled")

    def test_verified_outcome_with_running_remaining_blocks_start(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"RUNNING","bot_result":"working"}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"ADDED","service_result":"added"}],
        )
        app=make_fake_app(self.path)
        app.process=None
        payload={
            "outcome":"already_clean",
            "verified":True,
            "running_count":1,
            "job_ref":"A",
        }
        with mock.patch.object(tvc_control.messagebox,"showerror"):
            app._on_reconciliation(payload,1)
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")
        self.assertTrue(app.recovery_failed)
        self.assertEqual(app.start_button["state"],"disabled")

    def test_start_uses_validated_console_python_not_pythonw(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"WAIT","bot_result":""}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"WAIT","service_result":""}],
        )
        console_python=Path(self.tmp.name)/"python.exe"
        console_python.write_bytes(b"mock")
        app=make_fake_app(self.path)
        app.bot_python=console_python.resolve()
        process=FakeProcess(returncode=None)
        with (
            mock.patch.object(app,"_select_excel",return_value=True),
            mock.patch.object(tvc_control.subprocess,"Popen",return_value=process) as popen,
            mock.patch.object(tvc_control.threading,"Thread") as thread,
        ):
            app.start_bot()
        command=popen.call_args.args[0]
        self.assertEqual(Path(command[0]).name.lower(),"python.exe")
        self.assertNotIn("pythonw.exe",command[0].lower())
        self.assertEqual(command[1],"-u")
        self.assertEqual(thread.call_count,1)
        if app.stop_file is not None:
            app.stop_file.unlink(missing_ok=True)

    def test_invalid_runtime_blocks_bot_popen(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"WAIT","bot_result":""}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"WAIT","service_result":""}],
        )
        app=make_fake_app(self.path)
        app.runtime_valid=False
        app.bot_python=None
        with (
            mock.patch.object(tvc_control.subprocess,"Popen") as popen,
            mock.patch.object(tvc_control.messagebox,"showerror") as showerror,
        ):
            app._on_runtime_check_result(False,None,"broken venv")
            app.start_bot()
        popen.assert_not_called()
        showerror.assert_called_once()
        self.assertEqual(app.status_var.get(),"Error")
        self.assertEqual(app.start_button["state"],"disabled")
        self.assertEqual(app.runtime_button["state"],"normal")

    def test_runtime_check_is_non_blocking_and_success_enables_start(self):
        create_workbook(
            self.path,
            [{"job_ref":"A","bot_status":"WAIT","bot_result":""}],
            [{"job_ref":"A","service_seq":1,"service_code":"S","service_status":"WAIT","service_result":""}],
        )
        app=make_fake_app(self.path)
        app.runtime_valid=False
        app.bot_python=None
        started=threading.Event()
        release=threading.Event()
        console_python=Path(self.tmp.name)/"python.exe"
        console_python.write_bytes(b"mock")

        def slow_check():
            started.set()
            release.wait(2)
            return console_python.resolve()

        with mock.patch.object(tvc_control,"validate_bot_runtime",side_effect=slow_check):
            before=time.monotonic()
            app.retry_runtime_check()
            elapsed=time.monotonic()-before
            self.assertLess(elapsed,0.25)
            self.assertTrue(started.wait(1))
            self.assertEqual(app.status_var.get(),"กำลังตรวจสอบระบบ...")
            self.assertEqual(app.start_button["state"],"disabled")
            release.set()
            event=app.events.get(timeout=2)

        self.assertEqual(event[0],"runtime_check_result")
        app._on_runtime_check_result(event[1],event[2],event[3])
        self.assertTrue(app.runtime_valid)
        self.assertEqual(app.status_var.get(),"พร้อมใช้งาน")
        self.assertEqual(app.start_button["state"],"normal")

    def test_failed_runtime_check_can_retry_without_restarting_gui(self):
        app=make_fake_app(self.path)
        app.runtime_valid=False
        app.bot_python=None
        with mock.patch.object(tvc_control.messagebox,"showerror"):
            app._on_runtime_check_result(False,None,"first failure")
        self.assertEqual(app.runtime_button["state"],"normal")

        console_python=Path(self.tmp.name)/"python.exe"
        console_python.write_bytes(b"mock")
        with (
            mock.patch.object(tvc_control,"validate_bot_runtime",return_value=console_python.resolve()),
            mock.patch.object(tvc_control.threading,"Thread",ImmediateThread),
        ):
            app.retry_runtime_check()
        event=app.events.get_nowait()
        app._on_runtime_check_result(event[1],event[2],event[3])
        self.assertTrue(app.runtime_valid)
        self.assertEqual(app.start_button["state"],"normal")

    def test_console_dummy_subprocess_streams_log_and_progress_event(self):
        console_python=Path(sys.base_prefix)/"python.exe"
        if not console_python.is_file() or console_python.stat().st_size<=0:
            self.skipTest("ไม่มี console Python สำหรับ dummy subprocess")
        app=make_fake_app(self.path)
        code=(
            "print('DUMMY LOG',flush=True);"
            "print('TVCBOT_EVENT {\"event\":\"job_start\","
            "\"ref\":\"DUMMY-001\",\"index\":1,\"total\":2}',flush=True)"
        )
        process=subprocess.Popen(
            [str(console_python),"-u","-c",code],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        app.process=process
        app._read_process_output()
        events=[]
        while not app.events.empty():
            events.append(app.events.get_nowait())
        for event in events:
            if event[0]=="line":
                app._handle_line(event[1])
        self.assertIn("DUMMY LOG",app.logs)
        self.assertEqual(app.current_job_ref,"DUMMY-001")
        self.assertEqual(app.current_job_var.get(),"DUMMY-001 / 2")
        self.assertTrue(any(event[0]=="process_end" and event[2]==0 for event in events))


class BotMarkerTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(prefix="tvc_bot_marker_tests_")
        self.excel=Path(self.tmp.name)/"exists.xlsx"
        self.excel.write_bytes(b"fixture")
        self.stop_file=Path(self.tmp.name)/"stop.flag"
        self.jobs=[{
            "job_ref":"A",
            "_row":2,
            "plate_no":"ABC",
            "_services":[{
                "_row":2,
                "job_ref":"A",
                "service_seq":1,
                "service_code":"S",
            }],
        }]

    def tearDown(self):
        self.tmp.cleanup()

    def _run_bot(self,save_error=None):
        writes=[]
        service_writes=[]

        class Window:
            def window_text(self):
                return "MOCK"

        class Driver:
            def __init__(self,*args,**kwargs):
                self.count=0
            def connect(self):
                return Window()
            def set_text(self,*args,**kwargs):
                pass
            def set_date(self,*args,**kwargs):
                pass
            def list_count(self,*args,**kwargs):
                self.count+=1
                return 0 if self.count%2 else 1
            def click(self,*args,**kwargs):
                pass
            def screenshot(self,*args,**kwargs):
                pass
            def save_flow_yes_then_no(self,*args,**kwargs):
                if save_error:
                    raise save_error
                return {"first_clicked":"YES","second_clicked":"NO"}

        with (
            mock.patch.object(bot,"load_jobs",return_value=self.jobs),
            mock.patch.object(bot,"write_job_result",side_effect=lambda p,r,s,result="",tvc_job_no="":writes.append((s,result))),
            mock.patch.object(bot,"write_service_result",side_effect=lambda p,r,s,result="":service_writes.append((s,result))),
            mock.patch.object(bot,"TVCDriver",Driver),
            mock.patch.object(bot.time,"sleep"),
            mock.patch.object(bot.logging,"basicConfig"),
            mock.patch.object(bot.logging,"FileHandler",return_value=object()),
            mock.patch.object(bot.logging,"StreamHandler",return_value=object()),
            mock.patch.object(bot.logging,"exception"),
        ):
            rc=bot.main([
                "--excel",str(self.excel),
                "--stop-file",str(self.stop_file),
            ])
        return rc,writes,service_writes

    def test_commit_markers_surround_save(self):
        rc,writes,service_writes=self._run_bot()
        self.assertEqual(rc,0)
        self.assertEqual(
            writes[:4],
            [
                ("RUNNING","Bot กำลังทำงาน"),
                ("RUNNING",COMMITTING_TVC),
                ("RUNNING",TVC_SAVED_PENDING_EXCEL),
                ("DONE","บันทึกสำเร็จ รอบแรก=YES รอบสอง=NO"),
            ],
        )
        self.assertEqual(service_writes[0][0],"ADDED")

    def test_save_failure_is_marked_uncertain(self):
        rc,writes,_=self._run_bot(RuntimeError("mock save failure"))
        self.assertEqual(rc,1)
        self.assertEqual(writes[-1],("ERROR",UNCERTAIN_TVC_SAVE))

    def test_stop_requested_mid_job_finishes_current_safe_flow_only(self):
        jobs=[
            dict(self.jobs[0]),
            {
                **dict(self.jobs[0]),
                "job_ref":"B",
                "_row":3,
                "plate_no":"XYZ",
                "_services":[{
                    "_row":3,
                    "job_ref":"B",
                    "service_seq":1,
                    "service_code":"S2",
                }],
            },
        ]
        writes=[]
        saves=[]

        class Window:
            def window_text(self):
                return "MOCK"

        stop_file=self.stop_file

        class Driver:
            def __init__(self,*_args,**_kwargs):
                self.count=0
                self.stop_created=False
            def connect(self):
                return Window()
            def set_text(self,*_args,**_kwargs):
                if not self.stop_created:
                    stop_file.write_text("stop",encoding="utf-8")
                    self.stop_created=True
            def set_date(self,*_args,**_kwargs):
                pass
            def list_count(self,*_args,**_kwargs):
                self.count+=1
                return 0 if self.count%2 else 1
            def click(self,*_args,**_kwargs):
                pass
            def screenshot(self,*_args,**_kwargs):
                pass
            def save_flow_yes_then_no(self,*_args,**_kwargs):
                saves.append("save")
                return {"first_clicked":"YES","second_clicked":"NO"}

        with (
            mock.patch.object(bot,"load_jobs",return_value=jobs),
            mock.patch.object(
                bot,
                "write_job_result",
                side_effect=lambda _p,row,status,result="",tvc_job_no="":(
                    writes.append((row,status,result))
                ),
            ),
            mock.patch.object(bot,"write_service_result"),
            mock.patch.object(bot,"TVCDriver",Driver),
            mock.patch.object(bot.time,"sleep"),
            mock.patch.object(bot.logging,"basicConfig"),
            mock.patch.object(bot.logging,"FileHandler",return_value=object()),
            mock.patch.object(bot.logging,"StreamHandler",return_value=object()),
        ):
            rc=bot.main([
                "--excel",str(self.excel),
                "--stop-file",str(stop_file),
            ])

        self.assertEqual(rc,2)
        self.assertEqual(saves,["save"])
        self.assertTrue(any(row==2 and status=="DONE" for row,status,_ in writes))
        self.assertFalse(any(row==3 for row,_,_ in writes))


if __name__=="__main__":
    unittest.main(verbosity=2)
