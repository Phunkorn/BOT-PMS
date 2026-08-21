from pathlib import Path
import queue
import subprocess
import tempfile
import time
import tkinter as tk
import unittest
from unittest import mock

import sys
from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.table import Table


SRC=Path(__file__).resolve().parents[1]/"src"
sys.path.insert(0,str(SRC))

import tvc_control
import tvc_probe
import bot
from gui_queue import (
    EMPTY_STATS,
    ExcelQueue,
    ExcelQueueItem,
    QueueRunController,
    build_queue_summary,
    calculate_queue_progress,
)
from version import APP_NAME, APP_VERSION
from runtime_paths import build_worker_command,resolve_runtime_paths
from excel_io import (
    DIRTY_TVC_FORM_POSSIBLE,
    UNCERTAIN_TVC_SAVE,
    get_job_errors,
    get_job_stats,
    get_safety_issues,
    inspect_recovery_state,
    validate_workbook,
)


def stats(wait=0,done=0,error=0,running=0,other=0):
    total=wait+done+error+running+other
    return {
        "WAIT":wait,
        "DONE":done,
        "ERROR":error,
        "RUNNING":running,
        "OTHER":other,
        "TOTAL":total,
        "COMPLETED":done+error,
    }


class Value:
    def __init__(self,value=""):
        self.value=value
    def get(self):
        return self.value
    def set(self,value):
        self.value=value


def make_controller_app(paths):
    app=object.__new__(tvc_control.TVCControlApp)
    app.excel_queue=ExcelQueue()
    app.excel_queue.add_paths(paths)
    for item in app.excel_queue.items:
        item.status="READY"
        item.stats=stats(wait=1)
    app.queue_revision=1
    app.queue_running=False
    app.process=None
    app.precheck_valid=True
    app.precheck_generation=3
    app.precheck_in_progress=False
    app.precheck_purpose=""
    app.runtime_check_in_progress=False
    app.runtime_valid=True
    app.runtime_error=""
    app.bot_python=Path("python.exe").resolve()
    app.tvc_connected=True
    app.tvc_login_verified=True
    app.tvc_error=""
    app.start_pending=False
    app.finalization_in_progress=False
    app.recovery_in_progress=False
    app.recovery_failed=False
    app.closing=False
    app.safety_locks={}
    app.safety_metadata_health=tvc_control.SAFETY_METADATA_MISSING
    app.safety_metadata_error=""
    app.runtime_status_var=Value("Ready")
    app.tvc_status_var=Value("Connected")
    app.queue_status_var=Value("Ready")
    app.safety_status_var=Value("ไม่ล็อก")
    app.safety_detail_var=Value("")
    app.excel_var=Value(str(paths[0]) if paths else "")
    app.status_var=Value("")
    app._set_status=mock.Mock(side_effect=app.status_var.set)
    app._set_mascot=mock.Mock()
    app.append_log=mock.Mock()
    app._refresh_queue_tree=mock.Mock()
    app._refresh_progress=mock.Mock()
    app._update_buttons=mock.Mock()
    return app


def create_uncertain_workbook(path):
    wb=Workbook()
    jobs=wb.active
    jobs.title="JOB_INPUT"
    jobs.append(["job_ref","bot_status","bot_result"])
    jobs.append(["TEST-001","ERROR",UNCERTAIN_TVC_SAVE])
    jobs.add_table(Table(displayName="JobInputV5Table",ref="A1:C2"))
    services=wb.create_sheet("SERVICE_INPUT")
    services.append(
        ["job_ref","service_seq","service_code","service_status","service_result"]
    )
    services.append(["TEST-001",1,"SVC","ADDED","added"])
    services.add_table(Table(displayName="ServiceInputV5Table",ref="A1:E2"))
    wb.save(path)
    wb.close()


def create_uncertain_with_wait_workbook(path):
    wb=Workbook()
    jobs=wb.active
    jobs.title="JOB_INPUT"
    jobs.append(["job_ref","bot_status","bot_result"])
    jobs.append(["TEST-001","ERROR",UNCERTAIN_TVC_SAVE])
    jobs.append(["TEST-002","WAIT",""])
    jobs.add_table(Table(displayName="JobInputV5Table",ref="A1:C3"))
    services=wb.create_sheet("SERVICE_INPUT")
    services.append(
        ["job_ref","service_seq","service_code","service_status","service_result"]
    )
    services.append(["TEST-001",1,"SVC-1","ADDED","added"])
    services.append(["TEST-002",1,"SVC-2","WAIT",""])
    services.add_table(Table(displayName="ServiceInputV5Table",ref="A1:E3"))
    wb.save(path)
    wb.close()


class QueueModelTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(prefix="tvc_queue_v080_")
        self.base=Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def test_add_ten_duplicate_remove_reorder_and_clear(self):
        queue=ExcelQueue()
        paths=[self.base/f"jobs_{index}.xlsx" for index in range(10)]
        added,duplicates=queue.add_paths(paths)
        self.assertEqual(len(added),10)
        self.assertEqual(duplicates,[])

        added,duplicates=queue.add_paths([paths[0],paths[1].resolve()])
        self.assertEqual(added,[])
        self.assertEqual(len(duplicates),2)

        moved=queue.move(9,-1)
        self.assertEqual(moved,8)
        self.assertEqual(queue.items[8].path,paths[9].resolve())
        removed=queue.remove_indices([0,3])
        self.assertEqual(len(removed),2)
        self.assertEqual(len(queue.items),8)
        self.assertEqual(len(queue.clear()),8)
        self.assertEqual(queue.items,[])

    def test_queue_is_immutable_while_locked(self):
        queue=ExcelQueue()
        queue.add_paths([self.base/"a.xlsx"])
        queue.locked=True
        with self.assertRaises(RuntimeError):
            queue.add_paths([self.base/"b.xlsx"])
        with self.assertRaises(RuntimeError):
            queue.remove_indices([0])
        with self.assertRaises(RuntimeError):
            queue.clear()
        with self.assertRaises(RuntimeError):
            queue.move(0,1)


class QueueExecutionTests(unittest.TestCase):
    def test_stop_path_failure_rolls_back_queue_without_popen_or_next_file(self):
        paths=[Path("first.xlsx").resolve(),Path("second.xlsx").resolve()]
        app=make_controller_app(paths)
        app.root=mock.Mock()
        app.runtime_paths=mock.Mock()
        app.excel_queue.locked=True
        app.queue_running=True
        app.stop_request_sent=False
        app.stop_event_seen=False
        app.stop_event_phase=""
        app.last_batch_success=None
        app.force_stop_used=False
        app.force_retry_available=False
        app.last_return_code=None
        app.current_file_var=Value("- / 2")
        app.current_job_var=Value("- / 0")
        app.current_job_ref=""
        app.run_controller=QueueRunController(2)
        app.current_file_index=app.run_controller.start()
        for name in (
            "start_button","stop_button","browse_button","remove_button",
            "clear_button","up_button","down_button","runtime_button",
            "safety_button",
        ):
            setattr(app,name,mock.Mock())
        app._update_buttons=lambda:tvc_control.TVCControlApp._update_buttons(app)

        with (
            mock.patch.object(
                tvc_control,
                "prepare_worker_launch",
                side_effect=PermissionError("runtime mkdir denied"),
            ),
            mock.patch.object(tvc_control.subprocess,"Popen") as popen,
            mock.patch.object(tvc_control.messagebox,"showerror") as dialog,
        ):
            app._start_queue_item(0)

        popen.assert_not_called()
        self.assertFalse(app.queue_running)
        self.assertIsNone(app.process)
        self.assertIsNone(app.stop_file)
        self.assertFalse(app.excel_queue.locked)
        self.assertEqual(app.excel_queue.items[0].status,"ERROR")
        self.assertEqual(app.excel_queue.items[1].status,"READY")
        self.assertEqual(app.run_controller.outcome,"ERROR")
        self.assertEqual(app.status_var.get(),tvc_control.STOP_FILE_PREPARATION_ERROR)
        self.assertEqual(
            app.browse_button.configure.call_args.kwargs["state"],
            "normal",
        )
        self.assertEqual(
            app.stop_button.configure.call_args.kwargs["state"],
            "disabled",
        )
        self.assertIn("runtime mkdir denied",app.excel_queue.items[0].message)
        self.assertIn(
            "runtime mkdir denied",
            app.append_log.call_args.args[0],
        )
        dialog.assert_called_once()

    def test_stop_path_failure_rolls_back_legacy_start(self):
        app=make_controller_app([Path("legacy.xlsx").resolve()])
        app.root=mock.Mock()
        app.runtime_paths=mock.Mock()
        app.stop_file=None
        app.stop_request_sent=False
        app._select_excel=mock.Mock(return_value=True)

        with (
            mock.patch.object(
                tvc_control,
                "prepare_worker_launch",
                side_effect=PermissionError("legacy runtime denied"),
            ),
            mock.patch.object(tvc_control.subprocess,"Popen") as popen,
            mock.patch.object(tvc_control.messagebox,"showerror") as dialog,
        ):
            app._start_single_legacy()

        popen.assert_not_called()
        self.assertIsNone(app.process)
        self.assertIsNone(app.stop_file)
        self.assertEqual(app.status_var.get(),tvc_control.STOP_FILE_PREPARATION_ERROR)
        self.assertIn("legacy runtime denied",app.append_log.call_args.args[0])
        dialog.assert_called_once()

    def test_file_one_success_advances_to_file_two_and_all_success_finishes(self):
        controller=QueueRunController(2)
        self.assertEqual(controller.start(),0)
        self.assertEqual(controller.complete_current(0),1)
        self.assertIsNone(controller.complete_current(0))
        self.assertEqual(controller.outcome,"COMPLETE")

    def test_file_three_nonzero_never_starts_file_four(self):
        controller=QueueRunController(4)
        self.assertEqual(controller.start(),0)
        self.assertEqual(controller.complete_current(0),1)
        self.assertEqual(controller.complete_current(0),2)
        self.assertIsNone(controller.complete_current(1))
        self.assertEqual(controller.current_index,2)
        self.assertEqual(controller.outcome,"ERROR")

    def test_stop_during_file_prevents_next_file(self):
        controller=QueueRunController(3)
        controller.start()
        self.assertIsNone(controller.complete_current(0,stop_requested=True))
        self.assertEqual(controller.outcome,"STOPPED")

    def test_gui_success_handler_starts_exactly_next_item(self):
        app=object.__new__(tvc_control.TVCControlApp)
        app.excel_queue=ExcelQueue()
        app.excel_queue.add_paths([Path("one.xlsx"),Path("two.xlsx")])
        for item in app.excel_queue.items:
            item.status="READY"
            item.stats=stats(wait=1)
        app.queue_running=True
        app.current_file_index=0
        app.stop_request_sent=False
        app.run_controller=QueueRunController(2)
        app.run_controller.start()
        app._refresh_queue_tree=mock.Mock()
        app._refresh_progress=mock.Mock()
        app._start_queue_item=mock.Mock()
        app._finish_queue=mock.Mock()
        app.append_log=mock.Mock()

        app._on_queue_file_result(0,0,True,stats(done=1),([], ""))

        self.assertEqual(app.excel_queue.items[0].status,"DONE")
        app._start_queue_item.assert_called_once_with(1)
        app._finish_queue.assert_not_called()

    def test_gui_stop_handler_does_not_start_next_item(self):
        app=object.__new__(tvc_control.TVCControlApp)
        app.excel_queue=ExcelQueue()
        app.excel_queue.add_paths([Path("one.xlsx"),Path("two.xlsx")])
        app.queue_running=True
        app.current_file_index=0
        app.stop_request_sent=True
        app.run_controller=QueueRunController(2)
        app.run_controller.start()
        app._refresh_queue_tree=mock.Mock()
        app._refresh_progress=mock.Mock()
        app._start_queue_item=mock.Mock()
        app._finish_queue=mock.Mock()
        app.append_log=mock.Mock()

        app._on_queue_file_result(0,0,True,stats(done=1),([], ""))

        app._start_queue_item.assert_not_called()
        app._finish_queue.assert_called_once_with("STOPPED")

    def test_queue_nonzero_recovery_stops_at_file_three(self):
        app=object.__new__(tvc_control.TVCControlApp)
        app.excel_queue=ExcelQueue()
        app.excel_queue.add_paths(
            [Path("one.xlsx"),Path("two.xlsx"),Path("three.xlsx"),Path("four.xlsx")]
        )
        for item in app.excel_queue.items:
            item.status="READY"
            item.stats=stats(wait=1)
        app.excel_queue.items[2].status="RUNNING"
        app.queue_running=True
        app.current_file_index=2
        app.run_controller=QueueRunController(4)
        app.run_controller.start()
        app.run_controller.complete_current(0)
        app.run_controller.complete_current(0)
        app.recovery_in_progress=True
        app.recovery_failed=False
        app.current_job_ref="TEST-003"
        app.stop_event_seen=False
        app.force_stop_used=False
        app.last_batch_success=False
        app.closing=False
        app.append_log=mock.Mock()
        app._set_status=mock.Mock()
        app._refresh_queue_tree=mock.Mock()
        app._update_buttons=mock.Mock()
        app._start_queue_item=mock.Mock()
        payload={
            "outcome":"recovered",
            "verified":True,
            "running_count":0,
            "job_ref":"TEST-003",
            "job_reset":True,
            "services_reset":1,
            "message":"recovered",
        }
        with mock.patch.object(tvc_control.threading,"Thread") as thread:
            app._on_reconciliation(payload,1)
        self.assertEqual(app.excel_queue.items[2].status,"ERROR")
        self.assertEqual(app.run_controller.outcome,"ERROR")
        self.assertEqual(app.run_controller.current_index,2)
        app._start_queue_item.assert_not_called()
        thread.return_value.start.assert_called_once()


class ProgressTests(unittest.TestCase):
    def test_current_and_overall_progress_have_no_off_by_one(self):
        first=ExcelQueueItem("first.xlsx",status="RUNNING",stats=stats(wait=12,done=7,error=1))
        second=ExcelQueueItem("second.xlsx",status="READY",stats=stats(wait=5,done=4,error=1))
        result=calculate_queue_progress([first,second],0)
        self.assertEqual(result["current_processed"],8)
        self.assertEqual(result["current"]["TOTAL"],20)
        self.assertEqual(result["overall_processed"],13)
        self.assertEqual(result["overall"]["TOTAL"],30)
        self.assertEqual(result["overall"]["DONE"],11)
        self.assertEqual(result["overall"]["ERROR"],2)
        self.assertEqual(result["overall"]["WAIT"],17)


class PrecheckTests(unittest.TestCase):
    def setUp(self):
        self.paths=[Path(f"jobs_{index}.xlsx").resolve() for index in range(3)]
        self.runtime=Path("python.exe").resolve()

    def run_check(self,*,probe=None,stats_by_path=None,invalid=None,access_error=None):
        probe=probe or {"connected":True,"login_verified":True,"matches":["T.V.C Client"]}
        stats_by_path=stats_by_path or {str(path):stats(wait=1) for path in self.paths}
        invalid=set(invalid or [])

        def validate(path):
            if str(path) in invalid:
                raise RuntimeError("invalid schema")
            return True

        def access(path):
            if access_error and str(path)==str(access_error):
                raise PermissionError("file locked")
            return True

        return tvc_control.perform_precheck(
            self.paths,
            runtime_validator=lambda:self.runtime,
            tvc_checker=lambda _python:probe,
            workbook_validator=validate,
            access_checker=access,
            stats_reader=lambda path:stats_by_path[str(path)],
            errors_reader=lambda _path:[],
        )

    def test_valid_queue_is_ready(self):
        result=self.run_check()
        self.assertTrue(result["ready"])
        self.assertTrue(result["queue_ready"])
        self.assertEqual(result["total_wait"],3)
        self.assertTrue(all(item["status"]=="READY" for item in result["items"]))

    def test_tvc_absent_blocks_start(self):
        result=self.run_check(
            probe={"connected":False,"login_verified":False,"matches":[]}
        )
        self.assertFalse(result["ready"])
        self.assertIn("ไม่พบ T.V.C Client",result["tvc"]["message"])

    def test_runtime_failure_blocks_without_skipping_excel_validation(self):
        calls=[]
        result=tvc_control.perform_precheck(
            self.paths,
            runtime_validator=mock.Mock(side_effect=RuntimeError("broken runtime")),
            tvc_checker=mock.Mock(),
            workbook_validator=lambda path:calls.append(path),
            access_checker=lambda _path:True,
            stats_reader=lambda _path:stats(wait=1),
            errors_reader=lambda _path:[],
        )
        self.assertFalse(result["ready"])
        self.assertEqual(len(calls),len(self.paths))

    def test_running_job_invalid_schema_and_file_lock_are_invalid(self):
        stats_by_path={str(path):stats(wait=1) for path in self.paths}
        stats_by_path[str(self.paths[0])]=stats(running=1)
        result=self.run_check(
            stats_by_path=stats_by_path,
            invalid={str(self.paths[1])},
            access_error=self.paths[2],
        )
        self.assertFalse(result["queue_ready"])
        self.assertEqual(
            [item["status"] for item in result["items"]],
            ["INVALID","INVALID","INVALID"],
        )

    def test_queue_requires_at_least_one_wait(self):
        result=self.run_check(
            stats_by_path={str(path):stats(done=2) for path in self.paths}
        )
        self.assertFalse(result["queue_ready"])
        self.assertIn("WAIT",result["queue_message"])

    def test_combined_precheck_is_dispatched_to_worker_thread(self):
        app=object.__new__(tvc_control.TVCControlApp)
        app.excel_queue=ExcelQueue()
        app.excel_queue.add_paths([self.paths[0]])
        app.queue_running=False
        app.process=None
        app.recovery_in_progress=False
        app.precheck_in_progress=False
        app.closing=False
        app.precheck_generation=0
        app.runtime_check_in_progress=False
        app.precheck_valid=False
        app.valid_excel=False
        app.runtime_valid=False
        app.runtime_error=""
        app.bot_python=None
        app.tvc_connected=False
        app.tvc_login_verified=False
        app.runtime_status_var=mock.Mock()
        app.tvc_status_var=mock.Mock()
        app.queue_status_var=mock.Mock()
        app._refresh_queue_tree=mock.Mock()
        app._set_status=mock.Mock()
        app.append_log=mock.Mock()
        app._update_buttons=mock.Mock()
        with mock.patch.object(tvc_control.threading,"Thread") as thread:
            app.retry_precheck()
        self.assertTrue(app.precheck_in_progress)
        self.assertIs(thread.call_args.kwargs["target"].__func__,app._precheck_worker.__func__)
        thread.return_value.start.assert_called_once()


class PrecheckResponsivenessTests(unittest.TestCase):
    def setUp(self):
        self.path=Path("queue responsiveness.xlsx").resolve()

    @staticmethod
    def attach_buttons(app):
        app.stop_request_sent=False
        for name in (
            "start_button","stop_button","browse_button","remove_button",
            "clear_button","up_button","down_button","runtime_button",
            "safety_button",
        ):
            setattr(app,name,mock.Mock())

    def test_tvc_probe_subprocess_timeout_returns_explicit_timeout(self):
        def timed_out(*_args,**_kwargs):
            raise subprocess.TimeoutExpired(["probe"],0.01)

        result=tvc_control.check_tvc_client(
            tvc_control.RUNTIME_PATHS.source_python,
            timeout_seconds=0.01,
            runner=timed_out,
        )
        self.assertEqual(result["status"],"TIMEOUT")
        self.assertTrue(result["timed_out"])
        self.assertFalse(result["connected"])

    def test_real_probe_process_is_terminated_at_hard_timeout(self):
        with tempfile.TemporaryDirectory(prefix="tvc bounded probe ") as tmp:
            root=Path(tmp)/"App With Spaces"
            probe_script=root/"src"/"tvc_probe.py"
            probe_script.parent.mkdir(parents=True)
            probe_script.write_text(
                "import time\ntime.sleep(30)\n",encoding="utf-8"
            )
            paths=resolve_runtime_paths(
                frozen=False,
                module_file=root/"src"/"runtime_paths.py",
                writable_probe=lambda _path:True,
            )
            started=time.monotonic()
            result=tvc_control.check_tvc_client(
                Path(sys.executable),
                runtime_paths=paths,
                timeout_seconds=0.15,
            )
            elapsed=time.monotonic()-started
        self.assertEqual(result["status"],"TIMEOUT")
        self.assertTrue(result["timed_out"])
        self.assertLess(elapsed,2.0)

    def test_timeout_result_releases_precheck_state_and_queue_controls(self):
        app=make_controller_app([])
        self.attach_buttons(app)
        app.precheck_in_progress=True
        app.precheck_purpose="manual"
        app.runtime_check_in_progress=True
        app.precheck_valid=False
        app.valid_excel=False
        result={
            "runtime":{"ready":True,"python":str(Path("python.exe").resolve()),"message":"Ready"},
            "tvc":{
                "ready":False,"connected":False,"login_verified":False,
                "status":"TIMEOUT","timed_out":True,"duration_ms":4000,
                "message":"T.V.C probe หมดเวลา",
            },
            "items":[],"queue_ready":False,"queue_message":"ยังไม่ได้เลือก Excel",
            "total_wait":0,"ready":False,
        }
        app._on_precheck_result(app.precheck_generation,result)
        tvc_control.TVCControlApp._update_buttons(app)
        self.assertFalse(app.precheck_in_progress)
        self.assertFalse(app.runtime_check_in_progress)
        self.assertEqual(app.tvc_status_var.get(),"หมดเวลาตรวจสอบ")
        self.assertEqual(
            app.browse_button.configure.call_args.kwargs["state"],"normal"
        )
        self.assertEqual(
            app.start_button.configure.call_args.kwargs["state"],"disabled"
        )

    def test_queue_controls_enabled_while_manual_probe_running_or_not_found(self):
        for running in (True,False):
            with self.subTest(precheck_running=running):
                app=make_controller_app([self.path])
                self.attach_buttons(app)
                app.precheck_valid=False
                app.runtime_valid=True
                app.tvc_connected=False
                app.tvc_login_verified=False
                app.precheck_in_progress=running
                app.precheck_purpose="manual"
                app.runtime_check_in_progress=running
                tvc_control.TVCControlApp._update_buttons(app)
                for name in (
                    "browse_button","remove_button","clear_button",
                    "up_button","down_button",
                ):
                    button=getattr(app,name)
                    self.assertEqual(
                        button.configure.call_args.kwargs["state"],"normal",name
                    )
                self.assertEqual(
                    app.start_button.configure.call_args.kwargs["state"],"disabled"
                )

    def test_add_excel_during_manual_probe_invalidates_old_generation_and_retries(self):
        app=make_controller_app([])
        app.precheck_in_progress=True
        app.precheck_purpose="manual"
        app.runtime_check_in_progress=True
        old_generation=app.precheck_generation
        app.retry_precheck=mock.Mock()
        with mock.patch.object(
            tvc_control.filedialog,"askopenfilenames",return_value=(str(self.path),)
        ):
            app.choose_excel()
        self.assertEqual(len(app.excel_queue.items),1)
        self.assertGreater(app.precheck_generation,old_generation)
        self.assertFalse(app.precheck_in_progress)
        app.retry_precheck.assert_called_once_with()

    def test_clear_queue_during_manual_probe_retries_with_fresh_generation(self):
        app=make_controller_app([self.path])
        app.precheck_in_progress=True
        app.precheck_purpose="manual"
        app.runtime_check_in_progress=True
        old_generation=app.precheck_generation
        app.retry_precheck=mock.Mock()
        app.clear_files()
        self.assertEqual(app.excel_queue.items,[])
        self.assertGreater(app.precheck_generation,old_generation)
        self.assertFalse(app.precheck_in_progress)
        app.retry_precheck.assert_called_once_with()

    def test_stale_probe_result_cannot_overwrite_current_state(self):
        app=make_controller_app([])
        app.precheck_generation=9
        app.precheck_in_progress=True
        app.runtime_status_var.set("กำลังตรวจสอบ")
        app.tvc_status_var.set("กำลังตรวจสอบ")
        stale={
            "runtime":{"ready":True,"python":str(Path("python.exe").resolve()),"message":"Ready"},
            "tvc":{
                "ready":True,"connected":True,"login_verified":True,
                "status":"READY","message":"Connected",
            },
            "items":[],"queue_ready":False,"queue_message":"old",
            "total_wait":0,"ready":False,
        }
        app._on_precheck_result(8,stale)
        self.assertTrue(app.precheck_in_progress)
        self.assertEqual(app.tvc_status_var.get(),"กำลังตรวจสอบ")
        self.assertEqual(app.queue_status_var.get(),"Ready")

    def test_refresh_can_move_tvc_from_not_found_to_ready(self):
        app=make_controller_app([self.path])
        not_found={
            "runtime":{"ready":True,"python":str(Path("python.exe").resolve()),"message":"Ready"},
            "tvc":{
                "ready":False,"connected":False,"login_verified":False,
                "status":"NOT_FOUND","message":"ไม่พบ T.V.C Client",
            },
            "items":[{
                "path":str(self.path),"status":"READY","stats":stats(wait=1),
                "errors":[],"safety_issues":[],"message":"READY",
            }],
            "queue_ready":True,"queue_message":"Ready","total_wait":1,"ready":False,
        }
        app._on_precheck_result(app.precheck_generation,not_found)
        self.assertEqual(app.tvc_status_var.get(),"ยังไม่พบโปรแกรม")
        self.assertFalse(app.precheck_valid)

        app.precheck_generation+=1
        ready={**not_found,"ready":True,"tvc":{
            **not_found["tvc"],"ready":True,"connected":True,
            "login_verified":True,"status":"READY","message":"Connected",
        }}
        app._on_precheck_result(app.precheck_generation,ready)
        self.assertEqual(app.tvc_status_var.get(),"พร้อม")
        self.assertTrue(app.precheck_valid)

    def test_login_required_keeps_start_disabled_and_explains_state(self):
        app=make_controller_app([self.path])
        self.attach_buttons(app)
        result={
            "runtime":{"ready":True,"python":str(Path("python.exe").resolve()),"message":"Ready"},
            "tvc":{
                "ready":False,"connected":True,"login_verified":False,
                "status":"LOGIN_REQUIRED",
                "message":"พบ T.V.C แต่ยังยืนยันการเข้าสู่ระบบไม่ได้",
            },
            "items":[{
                "path":str(self.path),"status":"READY","stats":stats(wait=1),
                "errors":[],"safety_issues":[],"message":"READY",
            }],
            "queue_ready":True,"queue_message":"Ready","total_wait":1,"ready":False,
        }
        app._on_precheck_result(app.precheck_generation,result)
        tvc_control.TVCControlApp._update_buttons(app)
        self.assertEqual(
            app.tvc_status_var.get(),"พบโปรแกรม - ยังไม่ยืนยัน Login"
        )
        self.assertFalse(app.precheck_valid)
        self.assertEqual(
            app.start_button.configure.call_args.kwargs["state"],"disabled"
        )


class StartTimeRevalidationTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(prefix="tvc_start_revalidate_")
        self.path=Path(self.tmp.name)/"jobs.xlsx"
        self.path.write_bytes(b"fixture")
        self.runtime=Path("python.exe").resolve()

    def tearDown(self):
        self.tmp.cleanup()

    def fresh_result(self,workbook_validator=None,access_checker=None,job_stats=None):
        return tvc_control.perform_precheck(
            [self.path],
            runtime_validator=lambda:self.runtime,
            tvc_checker=lambda _python:{
                "connected":True,
                "login_verified":True,
                "matches":["T.V.C Client"],
            },
            workbook_validator=workbook_validator or (lambda _path:True),
            access_checker=access_checker or (lambda _path:True),
            stats_reader=lambda _path:job_stats or stats(wait=1),
            errors_reader=lambda _path:[],
        )

    def deliver_start_result(self,app,result,revision=1,path_keys=None):
        app.start_pending=True
        app.excel_queue.locked=True
        app.precheck_in_progress=True
        app.runtime_check_in_progress=True
        app._start_validated_queue=mock.Mock()
        app._on_precheck_result(
            app.precheck_generation,
            result,
            purpose="start",
            revision=revision,
            path_keys=path_keys or tuple(item.key for item in app.excel_queue.items),
        )
        return app._start_validated_queue

    def test_click_start_never_uses_cached_precheck_as_authorization(self):
        app=make_controller_app([self.path])
        app._begin_precheck=mock.Mock()
        with mock.patch.object(tvc_control.subprocess,"Popen") as popen:
            app.start_bot()
        popen.assert_not_called()
        self.assertTrue(app.start_pending)
        self.assertTrue(app.excel_queue.locked)
        app._begin_precheck.assert_called_once_with("start")

    def test_start_time_precheck_rechecks_tvc_and_blocks_if_it_closed(self):
        checker=mock.Mock(return_value={
            "status":"NOT_FOUND","connected":False,"login_verified":False,
            "matches":[],
        })
        result=tvc_control.perform_precheck(
            [self.path],
            runtime_validator=lambda:self.runtime,
            tvc_checker=checker,
            workbook_validator=lambda _path:True,
            access_checker=lambda _path:True,
            stats_reader=lambda _path:stats(wait=1),
            errors_reader=lambda _path:[],
        )
        checker.assert_called_once_with(self.runtime)
        self.assertFalse(result["tvc"]["ready"])
        self.assertFalse(result["ready"])

    def test_cached_pass_then_running_appears_blocks_start(self):
        app=make_controller_app([self.path])
        result=self.fresh_result(job_stats=stats(running=1))
        starter=self.deliver_start_result(app,result)
        starter.assert_not_called()
        self.assertFalse(app.excel_queue.locked)
        self.assertFalse(app.precheck_valid)

    def test_cached_pass_then_file_deleted_blocks_start(self):
        app=make_controller_app([self.path])
        result=self.fresh_result(
            workbook_validator=mock.Mock(side_effect=FileNotFoundError("deleted"))
        )
        starter=self.deliver_start_result(app,result)
        starter.assert_not_called()
        self.assertEqual(app.excel_queue.items[0].status,"INVALID")

    def test_cached_pass_then_file_locked_blocks_start(self):
        app=make_controller_app([self.path])
        result=self.fresh_result(
            access_checker=mock.Mock(side_effect=PermissionError("locked"))
        )
        starter=self.deliver_start_result(app,result)
        starter.assert_not_called()
        self.assertIn("locked",app.excel_queue.items[0].message)

    def test_queue_path_set_change_cancels_pending_start(self):
        app=make_controller_app([self.path])
        old_keys=tuple(item.key for item in app.excel_queue.items)
        result=self.fresh_result()
        app.excel_queue.add_paths([Path(self.tmp.name)/"other.xlsx"])
        starter=self.deliver_start_result(app,result,path_keys=old_keys)
        starter.assert_not_called()
        self.assertFalse(app.excel_queue.locked)
        self.assertFalse(app.start_pending)

    def test_fresh_start_revalidation_pass_auto_starts_queue(self):
        app=make_controller_app([self.path])
        starter=self.deliver_start_result(app,self.fresh_result())
        starter.assert_called_once_with()
        self.assertTrue(app.excel_queue.locked)


class FinalOutcomeTests(unittest.TestCase):
    def test_outcome_rules_cover_success_errors_stop_and_incomplete(self):
        success=build_queue_summary([
            ExcelQueueItem("done.xlsx",status="DONE",stats=stats(done=2))
        ])
        with_errors=build_queue_summary([
            ExcelQueueItem("error.xlsx",status="DONE",stats=stats(done=1,error=1))
        ])
        incomplete=build_queue_summary([
            ExcelQueueItem("wait.xlsx",status="DONE",stats=stats(done=1,wait=1))
        ])
        self.assertEqual(
            tvc_control.determine_final_outcome("COMPLETE",success),
            "COMPLETE_SUCCESS",
        )
        self.assertEqual(
            tvc_control.determine_final_outcome("COMPLETE",with_errors),
            "COMPLETE_WITH_ERRORS",
        )
        self.assertEqual(
            tvc_control.determine_final_outcome("ERROR",with_errors),
            "FAILED",
        )
        self.assertEqual(
            tvc_control.determine_final_outcome("STOPPED",incomplete),
            "STOPPED",
        )
        self.assertEqual(
            tvc_control.determine_final_outcome("COMPLETE",incomplete),
            "INCOMPLETE",
        )

    def test_final_handler_never_uses_success_status_when_error_exists(self):
        app=make_controller_app([Path("error.xlsx")])
        item=app.excel_queue.items[0]
        item.status="DONE"
        app.queue_running=True
        app.finalization_in_progress=True
        app.current_file_index=0
        app.process=None
        app.root=mock.Mock()
        result_stats=stats(done=1,error=1)
        with mock.patch.object(tvc_control.messagebox,"showwarning") as warning:
            app._on_queue_finalized(
                "COMPLETE",
                [(item.key,result_stats,[{"job_ref":"E-1","bot_result":"ERROR"}],"")],
            )
        statuses=[call.args[0] for call in app._set_status.call_args_list]
        self.assertNotIn("สำเร็จ",statuses)
        self.assertIn("เสร็จสิ้นพร้อมข้อผิดพลาด",statuses)
        warning.assert_called_once()

    def test_final_worker_refreshes_every_excel(self):
        app=object.__new__(tvc_control.TVCControlApp)
        app.events=queue.Queue()
        snapshot=[
            ("one",Path("one.xlsx")),
            ("two",Path("two.xlsx")),
            ("three",Path("three.xlsx")),
        ]
        with (
            mock.patch.object(tvc_control,"get_job_stats",return_value=stats(done=1)) as reader,
            mock.patch.object(tvc_control,"get_job_errors",return_value=[]),
        ):
            app._finalize_queue_worker("COMPLETE",snapshot)
        self.assertEqual(reader.call_count,3)
        event=app.events.get_nowait()
        self.assertEqual(event[0],"queue_finalized")
        self.assertEqual(len(event[2]),3)


class SafetyLockTests(unittest.TestCase):
    def ready_precheck(self,paths):
        return {
            "runtime":{"ready":True,"python":str(Path("python.exe").resolve()),"message":"Ready"},
            "tvc":{"ready":True,"connected":True,"login_verified":True,"message":"Connected"},
            "items":[
                {
                    "path":str(Path(path).resolve()),
                    "status":"READY",
                    "stats":stats(wait=1),
                    "errors":[],
                    "message":"READY",
                }
                for path in paths
            ],
            "queue_ready":bool(paths),
            "queue_message":"Ready",
            "total_wait":len(paths),
            "ready":bool(paths),
        }

    def test_uncertain_commit_sets_safety_lock(self):
        path=Path("uncertain.xlsx").resolve()
        app=make_controller_app([path])
        app.current_job_ref="TEST-001"
        app.stop_event_seen=False
        app.force_stop_used=False
        app.last_batch_success=False
        app._request_stats=mock.Mock()
        app.root=mock.Mock()
        payload={
            "outcome":"uncertain_commit",
            "verified":True,
            "running_count":0,
            "job_ref":"TEST-001",
            "previous_status":"ERROR",
            "previous_result":UNCERTAIN_TVC_SAVE,
            "message":UNCERTAIN_TVC_SAVE,
        }
        with mock.patch.object(tvc_control.messagebox,"showerror"):
            app._on_reconciliation(payload,1)
        self.assertTrue(app._safety_lock_active())
        self.assertTrue(app.recovery_failed)
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")

    def test_precheck_pass_does_not_show_ready_while_lock_active(self):
        path=Path("locked.xlsx").resolve()
        app=make_controller_app([path])
        app.safety_locks={
            app.excel_queue.items[0].key:{
                "path":str(path),
                "outcome":"uncertain_commit",
                "message":UNCERTAIN_TVC_SAVE,
                "job_ref":"TEST-001",
            }
        }
        app.recovery_failed=True
        app._on_precheck_result(
            app.precheck_generation,
            self.ready_precheck([path]),
            purpose="manual",
            revision=app.queue_revision,
            path_keys=tuple(item.key for item in app.excel_queue.items),
        )
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")
        self.assertNotEqual(app.status_var.get(),"พร้อมใช้งาน")

    def test_safety_lock_disables_start_and_enables_resolution_action(self):
        path=Path("locked.xlsx").resolve()
        app=make_controller_app([path])
        app.safety_locks={
            app.excel_queue.items[0].key:{
                "path":str(path),
                "outcome":"uncertain_commit",
                "message":UNCERTAIN_TVC_SAVE,
                "job_ref":"TEST-001",
            }
        }
        app.recovery_failed=True
        app.stop_request_sent=False
        for name in (
            "start_button","stop_button","browse_button","remove_button",
            "clear_button","up_button","down_button","runtime_button",
            "safety_button",
        ):
            setattr(app,name,mock.Mock())

        tvc_control.TVCControlApp._update_buttons(app)

        self.assertEqual(
            app.start_button.configure.call_args.kwargs["state"],"disabled"
        )
        self.assertEqual(
            app.safety_button.configure.call_args.kwargs["state"],"normal"
        )

    def test_running_still_present_keeps_lock(self):
        key=str(Path("locked.xlsx").resolve()).lower()
        locks={key:{"message":"ambiguous"}}
        precheck=self.ready_precheck([Path(key)])
        precheck["items"][0]["status"]="INVALID"
        precheck["ready"]=False
        resolved,_=tvc_control.evaluate_safety_revalidation(
            precheck,
            locks,
            {key:{"outcome":"pending_recovery","running_count":1,"verified":False}},
            [key],
        )
        self.assertFalse(resolved)

    def test_unresolved_uncertain_error_keeps_lock_and_never_changes_to_wait(self):
        with tempfile.TemporaryDirectory(prefix="tvc_uncertain_readonly_") as tmp:
            path=Path(tmp)/"uncertain.xlsx"
            create_uncertain_workbook(path)
            inspection=inspect_recovery_state(path,current_job_ref="TEST-001")
            key=str(path.resolve()).lower()
            locks={key:{"message":UNCERTAIN_TVC_SAVE}}
            resolved,_=tvc_control.evaluate_safety_revalidation(
                self.ready_precheck([path]),locks,{key:inspection},[key]
            )
            self.assertFalse(resolved)
            wb=load_workbook(path,data_only=False)
            try:
                self.assertEqual(wb["JOB_INPUT"].cell(2,2).value,"ERROR")
                self.assertEqual(wb["JOB_INPUT"].cell(2,3).value,UNCERTAIN_TVC_SAVE)
            finally:
                wb.close()

    def test_removed_problem_file_stays_unresolved_without_inspection(self):
        removed_key=str(Path("removed.xlsx").resolve()).lower()
        precheck={
            "runtime":{"ready":True},
            "tvc":{"ready":True},
            "items":[],
            "ready":False,
        }
        resolved,unresolved=tvc_control.evaluate_safety_revalidation(
            precheck,
            {removed_key:{"message":"failed"}},
            {},
            [],
        )
        self.assertFalse(resolved)
        self.assertIn("ไม่มีผลตรวจ recovery",unresolved[0])

    def test_explicit_revalidation_clears_lock_without_restart(self):
        path=Path("fixed.xlsx").resolve()
        app=make_controller_app([path])
        key=app.excel_queue.items[0].key
        app.safety_locks={
            key:{
                "path":str(path),
                "outcome":"ambiguous",
                "message":"fixed manually",
                "job_ref":"TEST-001",
            }
        }
        app.recovery_failed=True
        app._on_safety_revalidation_result(
            app.precheck_generation,
            app.queue_revision,
            (key,),
            self.ready_precheck([path]),
            True,
            [],
        )
        self.assertFalse(app._safety_lock_active())
        self.assertFalse(app.recovery_failed)
        self.assertEqual(app.status_var.get(),"พร้อมใช้งาน")

    def test_dirty_form_lock_requires_tvc_job_form_to_be_closed(self):
        path=Path("dirty.xlsx").resolve()
        key=str(path).lower()
        locks={key:{"outcome":"DIRTY_TVC_FORM_POSSIBLE","message":"dirty"}}
        inspection={
            key:{"outcome":"already_clean","verified":True,"running_count":0}
        }
        precheck=self.ready_precheck([path])
        precheck["tvc"]["active_job_form"]=True
        resolved,_=tvc_control.evaluate_safety_revalidation(
            precheck,locks,inspection,[key]
        )
        self.assertFalse(resolved)

        precheck["tvc"]["active_job_form"]=False
        resolved,_=tvc_control.evaluate_safety_revalidation(
            precheck,locks,inspection,[key]
        )
        self.assertTrue(resolved)

    def test_recovered_precommit_creates_dirty_form_lock_and_blocks_start(self):
        path=Path("dirty.xlsx").resolve()
        app=make_controller_app([path])
        app.current_job_ref="TEST-001"
        app.stop_event_seen=False
        app.force_stop_used=True
        app.last_batch_success=False
        app._request_stats=mock.Mock()
        payload={
            "outcome":"recovered",
            "verified":True,
            "running_count":0,
            "job_ref":"TEST-001",
            "job_reset":True,
            "services_reset":1,
            "message":DIRTY_TVC_FORM_POSSIBLE,
        }
        app._on_reconciliation(payload,1)
        self.assertTrue(app._safety_lock_active())
        lock=next(iter(app.safety_locks.values()))
        self.assertEqual(lock["outcome"],"DIRTY_TVC_FORM_POSSIBLE")
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")

        app._begin_precheck=mock.Mock()
        app.start_bot()
        app._begin_precheck.assert_not_called()

    def test_dirty_lock_releases_only_after_explicit_clean_revalidation(self):
        path=Path("dirty fixed.xlsx").resolve()
        app=make_controller_app([path])
        key=app.excel_queue.items[0].key
        app.safety_locks={
            key:{
                "path":str(path),
                "outcome":"DIRTY_TVC_FORM_POSSIBLE",
                "message":"dirty",
                "job_ref":"TEST-001",
            }
        }
        app.recovery_failed=True
        precheck=self.ready_precheck([path])
        precheck["tvc"]["active_job_form"]=False
        app._on_safety_revalidation_result(
            app.precheck_generation,
            app.queue_revision,
            (key,),
            precheck,
            True,
            [],
        )
        self.assertFalse(app._safety_lock_active())
        self.assertFalse(app.recovery_failed)
        self.assertEqual(app.status_var.get(),"พร้อมใช้งาน")

    def test_dirty_lock_metadata_survives_gui_restart_until_reviewed(self):
        with tempfile.TemporaryDirectory(prefix="tvc safety state ") as tmp:
            state_file=Path(tmp)/"safety_locks.json"
            path=Path(tmp)/"jobs.xlsx"
            key=str(path.resolve()).lower()
            locks={
                key:{
                    "path":str(path),
                    "outcome":"DIRTY_TVC_FORM_POSSIBLE",
                    "message":"dirty",
                    "job_ref":"TEST-001",
                }
            }
            tvc_control.save_persisted_safety_locks(state_file,locks)
            restored=tvc_control.load_persisted_safety_locks(state_file)
            self.assertEqual(restored.health,tvc_control.SAFETY_METADATA_HEALTHY)
            self.assertIn(key,restored.locks)
            self.assertEqual(
                restored.locks[key]["outcome"],"DIRTY_TVC_FORM_POSSIBLE"
            )
            tvc_control.save_persisted_safety_locks(state_file,{})
            self.assertFalse(state_file.exists())


class SafetyMetadataFailureTests(unittest.TestCase):
    def test_missing_and_healthy_metadata_have_explicit_health(self):
        with tempfile.TemporaryDirectory(prefix="tvc metadata health ") as tmp:
            state_file=Path(tmp)/"safety_locks.json"
            missing=tvc_control.load_persisted_safety_locks(state_file)
            self.assertEqual(missing.health,tvc_control.SAFETY_METADATA_MISSING)
            self.assertEqual(missing.locks,{})

            workbook=Path(tmp)/"jobs.xlsx"
            key=str(workbook.resolve()).lower()
            locks={key:{
                "path":str(workbook.resolve()),
                "outcome":"DIRTY_TVC_FORM_POSSIBLE",
                "message":"dirty",
                "job_ref":"TEST-001",
            }}
            tvc_control.save_persisted_safety_locks(state_file,locks)
            healthy=tvc_control.load_persisted_safety_locks(state_file)
            self.assertEqual(healthy.health,tvc_control.SAFETY_METADATA_HEALTHY)
            self.assertEqual(healthy.locks,locks)

    def test_corrupt_json_is_fail_closed_and_not_overwritten_automatically(self):
        with tempfile.TemporaryDirectory(prefix="tvc corrupt metadata ") as tmp:
            state_file=Path(tmp)/"safety_locks.json"
            original="{ definitely not json"
            state_file.write_text(original,encoding="utf-8")
            loaded=tvc_control.load_persisted_safety_locks(state_file)
            self.assertEqual(loaded.health,tvc_control.SAFETY_METADATA_CORRUPT)
            self.assertEqual(loaded.locks,{})

            app=make_controller_app([])
            app.safety_state_file=state_file
            app.safety_metadata_health=loaded.health
            app.safety_metadata_error=loaded.message
            with mock.patch.object(tvc_control,"save_persisted_safety_locks") as save:
                self.assertFalse(app._persist_safety_locks())
            save.assert_not_called()
            self.assertEqual(state_file.read_text(encoding="utf-8"),original)
            self.assertTrue(app._safety_lock_active())

    def test_invalid_utf8_metadata_is_corrupt_not_startup_exception(self):
        with tempfile.TemporaryDirectory(prefix="tvc invalid utf8 metadata ") as tmp:
            state_file=Path(tmp)/"safety_locks.json"
            state_file.write_bytes(b"\xff\xfe\xfa")
            loaded=tvc_control.load_persisted_safety_locks(state_file)
            self.assertEqual(loaded.health,tvc_control.SAFETY_METADATA_CORRUPT)
            self.assertEqual(loaded.locks,{})
            self.assertIn("encoding",loaded.message)

    def test_unreadable_and_permission_denied_metadata_are_explicit(self):
        with tempfile.TemporaryDirectory(prefix="tvc unreadable metadata ") as tmp:
            state_file=Path(tmp)/"safety_locks.json"
            state_file.write_text("{}",encoding="utf-8")
            for error in (PermissionError("denied"),OSError("unreadable")):
                with self.subTest(error=type(error).__name__):
                    with mock.patch.object(Path,"read_text",side_effect=error):
                        loaded=tvc_control.load_persisted_safety_locks(state_file)
                    self.assertEqual(
                        loaded.health,tvc_control.SAFETY_METADATA_UNREADABLE
                    )
                    self.assertTrue(loaded.message)
            with mock.patch.object(Path,"stat",side_effect=PermissionError("stat denied")):
                loaded=tvc_control.load_persisted_safety_locks(state_file)
            self.assertEqual(loaded.health,tvc_control.SAFETY_METADATA_UNREADABLE)

    def test_write_fsync_and_atomic_replace_failures_raise(self):
        with tempfile.TemporaryDirectory(prefix="tvc metadata writes ") as tmp:
            state_file=Path(tmp)/"safety_locks.json"
            locks={"x":{"path":str(Path(tmp)/"x.xlsx"),"outcome":"dirty"}}
            with mock.patch.object(Path,"open",side_effect=OSError("disk full")):
                with self.assertRaisesRegex(OSError,"disk full"):
                    tvc_control.save_persisted_safety_locks(state_file,locks)
            with mock.patch.object(tvc_control.os,"fsync",side_effect=OSError("fsync failed")):
                with self.assertRaisesRegex(OSError,"fsync failed"):
                    tvc_control.save_persisted_safety_locks(state_file,locks)

            state_file.write_text('{"old": true}',encoding="utf-8")
            with mock.patch.object(Path,"replace",side_effect=PermissionError("replace denied")):
                with self.assertRaisesRegex(PermissionError,"replace denied"):
                    tvc_control.save_persisted_safety_locks(state_file,locks)
            self.assertEqual(state_file.read_text(encoding="utf-8"),'{"old": true}')
            self.assertFalse(state_file.with_suffix(".json.tmp").exists())

    def test_write_failure_keeps_memory_lock_and_disables_start(self):
        with tempfile.TemporaryDirectory(prefix="tvc write fail gate ") as tmp:
            workbook=Path(tmp)/"jobs.xlsx"
            app=make_controller_app([workbook])
            key=app.excel_queue.items[0].key
            app.safety_state_file=Path(tmp)/"safety_locks.json"
            app.safety_locks={key:{
                "path":str(workbook.resolve()),
                "outcome":"DIRTY_TVC_FORM_POSSIBLE",
                "message":"dirty",
                "job_ref":"TEST-001",
            }}
            with mock.patch.object(
                tvc_control,"save_persisted_safety_locks",side_effect=PermissionError("denied")
            ):
                self.assertFalse(app._persist_safety_locks())
            self.assertIn(key,app.safety_locks)
            self.assertEqual(
                app.safety_metadata_health,tvc_control.SAFETY_METADATA_WRITE_FAILED
            )
            self.assertEqual(app.status_var.get(),tvc_control.SAFETY_PERSISTENCE_ERROR)
            self.assertTrue(app.recovery_failed)
            self.assertTrue(app._retry_safety_metadata())
            self.assertEqual(
                app.safety_metadata_health,tvc_control.SAFETY_METADATA_HEALTHY
            )
            self.assertIn(
                key,
                tvc_control.load_persisted_safety_locks(app.safety_state_file).locks,
            )

            app.stop_request_sent=False
            for name in (
                "start_button","stop_button","browse_button","remove_button",
                "clear_button","up_button","down_button","runtime_button",
                "safety_button",
            ):
                setattr(app,name,mock.Mock())
            tvc_control.TVCControlApp._update_buttons(app)
            self.assertEqual(
                app.start_button.configure.call_args.kwargs["state"],"disabled"
            )

    def test_restart_with_corrupt_metadata_is_fail_closed_and_retry_can_reload(self):
        with tempfile.TemporaryDirectory(prefix="tvc restart corrupt ") as tmp:
            state_file=Path(tmp)/"safety_locks.json"
            state_file.write_text("broken",encoding="utf-8")
            loaded=tvc_control.load_persisted_safety_locks(state_file)
            app=make_controller_app([])
            app.safety_state_file=state_file
            app.safety_locks=loaded.locks
            app.safety_metadata_health=loaded.health
            app.safety_metadata_error=loaded.message
            app.recovery_failed=True
            self.assertTrue(app._safety_lock_active())

            state_file.write_text("{}",encoding="utf-8")
            self.assertTrue(app._retry_safety_metadata())
            self.assertEqual(
                app.safety_metadata_health,tvc_control.SAFETY_METADATA_HEALTHY
            )
            self.assertFalse(app._safety_lock_active())

    def test_start_is_disabled_for_every_fail_closed_health_state(self):
        for health in (
            tvc_control.SAFETY_METADATA_CORRUPT,
            tvc_control.SAFETY_METADATA_UNREADABLE,
            tvc_control.SAFETY_METADATA_WRITE_FAILED,
        ):
            with self.subTest(health=health):
                app=make_controller_app([Path(f"{health}.xlsx").resolve()])
                app.safety_metadata_health=health
                app.safety_metadata_error="metadata failure"
                app.recovery_failed=True
                app.stop_request_sent=False
                for name in (
                    "start_button","stop_button","browse_button","remove_button",
                    "clear_button","up_button","down_button","runtime_button",
                    "safety_button",
                ):
                    setattr(app,name,mock.Mock())
                tvc_control.TVCControlApp._update_buttons(app)
                self.assertEqual(
                    app.start_button.configure.call_args.kwargs["state"],"disabled"
                )
                self.assertEqual(
                    app.safety_button.configure.call_args.kwargs["state"],"normal"
                )


class SafetyRegistryTests(unittest.TestCase):
    @staticmethod
    def ready_precheck(paths):
        return {
            "runtime":{"ready":True,"python":str(Path("python.exe").resolve()),"message":"Ready"},
            "tvc":{"ready":True,"connected":True,"login_verified":True,
                   "active_job_form":False,"message":"Connected"},
            "items":[{
                "path":str(path.resolve()),"status":"READY","stats":stats(wait=1),
                "errors":[],"safety_issues":[],"message":"READY",
            } for path in paths],
            "queue_ready":bool(paths),"queue_message":"Ready","total_wait":len(paths),
            "ready":bool(paths),
        }

    def test_removed_lock_is_independent_and_readd_uses_same_key(self):
        path=Path("removed dirty.xlsx").resolve()
        app=make_controller_app([path])
        key=app.excel_queue.items[0].key
        app.safety_locks={key:{
            "path":str(path),"outcome":"DIRTY_TVC_FORM_POSSIBLE",
            "message":"dirty","job_ref":"TEST-001",
        }}
        app.excel_queue.remove_indices([0])
        self.assertIn(key,app.safety_locks)
        self.assertTrue(app._safety_lock_active())
        app.excel_queue.add_paths([path])
        self.assertEqual(app.excel_queue.items[0].key,key)
        self.assertIn(key,app.safety_locks)

    def test_queue_empty_with_unresolved_lock_disables_start_and_lists_path(self):
        path=Path("unresolved outside queue.xlsx").resolve()
        key=str(path).lower()
        app=make_controller_app([])
        app.safety_locks={key:{
            "path":str(path),"outcome":"DIRTY_TVC_FORM_POSSIBLE",
            "message":"dirty","job_ref":"TEST-001",
        }}
        app.recovery_failed=True
        app.stop_request_sent=False
        for name in (
            "start_button","stop_button","browse_button","remove_button",
            "clear_button","up_button","down_button","runtime_button",
            "safety_button",
        ):
            setattr(app,name,mock.Mock())
        app._refresh_safety_status()
        tvc_control.TVCControlApp._update_buttons(app)
        self.assertEqual(
            app.start_button.configure.call_args.kwargs["state"],"disabled"
        )
        self.assertIn(str(path),app.safety_detail_var.get())
        self.assertIn("DIRTY_TVC_FORM_POSSIBLE",app.safety_detail_var.get())

    def test_safety_worker_inspects_lock_removed_from_queue(self):
        path=Path("removed but inspected.xlsx").resolve()
        key=str(path).lower()
        locks={key:{
            "path":str(path),"outcome":"DIRTY_TVC_FORM_POSSIBLE",
            "message":"dirty","job_ref":"TEST-001",
        }}
        app=make_controller_app([])
        app.events=queue.Queue()
        inspection={"outcome":"already_clean","verified":True,"running_count":0}
        with (
            mock.patch.object(tvc_control,"perform_precheck",return_value=self.ready_precheck([])),
            mock.patch.object(tvc_control,"inspect_recovery_state",return_value=inspection) as inspect,
            mock.patch.object(tvc_control,"get_safety_issues",return_value=[]),
        ):
            app._safety_validation_worker(1,2,[],(),locks)
        inspect.assert_called_once_with(path,current_job_ref="TEST-001")
        event=app.events.get_nowait()
        self.assertEqual(event[0],"safety_revalidation_result")
        self.assertTrue(event[5])
        self.assertEqual(event[7],(key,))

    def test_missing_lock_file_is_unresolved(self):
        path=Path("missing locked.xlsx").resolve()
        key=str(path).lower()
        resolved,unresolved,resolved_keys=tvc_control.evaluate_safety_revalidation(
            self.ready_precheck([]),
            {key:{"path":str(path),"outcome":"DIRTY_TVC_FORM_POSSIBLE"}},
            {key:{"outcome":"failed","verified":False,"running_count":-1,
                  "message":"ไม่พบไฟล์ Excel"}},
            [],
            include_resolved_keys=True,
        )
        self.assertFalse(resolved)
        self.assertEqual(resolved_keys,())
        self.assertIn("ไม่พบไฟล์",unresolved[0])

    def test_multiple_locks_release_only_clean_file(self):
        first=Path("first clean.xlsx").resolve()
        second=Path("second locked.xlsx").resolve()
        first_key=str(first).lower()
        second_key=str(second).lower()
        locks={
            first_key:{"path":str(first),"outcome":"DIRTY_TVC_FORM_POSSIBLE"},
            second_key:{"path":str(second),"outcome":"DIRTY_TVC_FORM_POSSIBLE"},
        }
        inspections={
            first_key:{"outcome":"already_clean","verified":True,"running_count":0,
                       "safety_issues":[]},
            second_key:{"outcome":"failed","verified":False,"running_count":-1,
                        "message":"unreadable","safety_issues":[]},
        }
        resolved,unresolved,resolved_keys=tvc_control.evaluate_safety_revalidation(
            self.ready_precheck([]),locks,inspections,[],include_resolved_keys=True
        )
        self.assertFalse(resolved)
        self.assertEqual(resolved_keys,(first_key,))
        self.assertIn("unreadable",unresolved[0])

        with tempfile.TemporaryDirectory(prefix="tvc partial release ") as tmp:
            app=make_controller_app([])
            app.safety_locks=dict(locks)
            app.safety_state_file=Path(tmp)/"safety_locks.json"
            app._on_safety_revalidation_result(
                app.precheck_generation,app.queue_revision,(),self.ready_precheck([]),
                resolved,unresolved,resolved_keys,
            )
            self.assertNotIn(first_key,app.safety_locks)
            self.assertIn(second_key,app.safety_locks)
            persisted=tvc_control.load_persisted_safety_locks(app.safety_state_file)
            self.assertEqual(set(persisted.locks),{second_key})


class PersistedSafetyLockTests(unittest.TestCase):
    def setUp(self):
        self.tmp=tempfile.TemporaryDirectory(prefix="tvc_persisted_lock_")
        self.path=Path(self.tmp.name)/"uncertain and wait.xlsx"
        create_uncertain_with_wait_workbook(self.path)

    def tearDown(self):
        self.tmp.cleanup()

    def precheck(self):
        return tvc_control.perform_precheck(
            [self.path],
            runtime_validator=lambda:Path("python.exe").resolve(),
            tvc_checker=lambda _:{
                "connected":True,
                "login_verified":True,
                "active_job_form":False,
                "matches":["T.V.C Client"],
            },
            workbook_validator=validate_workbook,
            access_checker=lambda _:True,
            stats_reader=get_job_stats,
            errors_reader=get_job_errors,
            safety_reader=get_safety_issues,
        )

    def apply_as_fresh_gui(self,result):
        app=make_controller_app([self.path])
        app.safety_locks={}
        app.recovery_failed=False
        app._on_precheck_result(
            app.precheck_generation,
            result,
            purpose="manual",
            revision=app.queue_revision,
            path_keys=tuple(item.key for item in app.excel_queue.items),
        )
        return app

    def test_fresh_startup_uncertain_plus_wait_is_locked_and_not_ready(self):
        result=self.precheck()
        self.assertFalse(result["ready"])
        self.assertFalse(result["queue_ready"])
        self.assertEqual(result["items"][0]["status"],"REVIEW_REQUIRED")

        app=self.apply_as_fresh_gui(result)
        self.assertTrue(app._safety_lock_active())
        self.assertEqual(app.status_var.get(),"ต้องตรวจสอบงานก่อนเริ่มใหม่")
        self.assertIn("ไม่ยืนยันผลการบันทึก",app.excel_queue.items[0].message)

        app.stop_request_sent=False
        for name in (
            "start_button","stop_button","browse_button","remove_button",
            "clear_button","up_button","down_button","runtime_button",
            "safety_button",
        ):
            setattr(app,name,mock.Mock())
        tvc_control.TVCControlApp._update_buttons(app)
        self.assertEqual(
            app.start_button.configure.call_args.kwargs["state"],"disabled"
        )

    def test_new_gui_instance_rebuilds_lock_from_same_workbook(self):
        first=self.apply_as_fresh_gui(self.precheck())
        second=self.apply_as_fresh_gui(self.precheck())
        self.assertTrue(first._safety_lock_active())
        self.assertTrue(second._safety_lock_active())
        self.assertIsNot(first.safety_locks,second.safety_locks)

        wb=load_workbook(self.path,data_only=False)
        try:
            self.assertEqual(wb["JOB_INPUT"].cell(2,2).value,"ERROR")
            self.assertEqual(wb["JOB_INPUT"].cell(2,3).value,UNCERTAIN_TVC_SAVE)
        finally:
            wb.close()


class RuntimePathTests(unittest.TestCase):
    def test_source_and_frozen_worker_commands_support_spaces(self):
        with tempfile.TemporaryDirectory(prefix="tvc runtime paths ") as tmp:
            root=Path(tmp)/"Project With Spaces"
            (root/"src").mkdir(parents=True)
            source=resolve_runtime_paths(
                frozen=False,module_file=root/"src"/"runtime_paths.py"
            )
            excel=root/"input files"/"jobs one.xlsx"
            stop=root/"temp files"/"stop one.flag"
            source_command=build_worker_command(source,excel,stop)
            self.assertEqual(source_command[0],str(root/".venv"/"Scripts"/"python.exe"))
            self.assertEqual(source_command[1:3],["-u",str(root/"src"/"bot.py")])
            self.assertIn(str(excel),source_command)
            self.assertEqual(source.config_file,root/"config.ini")
            self.assertEqual(source.assets_dir,root/"assets")

            app_dir=root/"dist folder"
            app_dir.mkdir(parents=True)
            frozen=resolve_runtime_paths(
                frozen=True,
                executable=app_dir/"TVC Bot Control.exe",
                bundle_dir=root/"bundle resources",
            )
            frozen_command=build_worker_command(frozen,excel,stop)
            self.assertEqual(frozen_command[0],str(app_dir/"TVC Bot Worker.exe"))
            self.assertNotIn(".venv"," ".join(frozen_command).lower())
            self.assertNotIn("bot.py"," ".join(frozen_command).lower())
            self.assertIn(str(excel),frozen_command)
            self.assertEqual(frozen.logs_dir,app_dir/"logs")

    def test_frozen_gui_worker_and_probe_resolve_same_external_config(self):
        with tempfile.TemporaryDirectory(prefix="tvc shared frozen config ") as tmp:
            app_dir=Path(tmp)/"Installed App With Spaces"
            app_dir.mkdir(parents=True)
            external=app_dir/"config.ini"
            external.write_text("[tvc]\nbackend = win32\n",encoding="utf-8")
            gui_paths=resolve_runtime_paths(
                frozen=True,
                executable=app_dir/"TVC Bot Control.exe",
                bundle_dir=app_dir/"gui bundle",
            )
            worker_paths=resolve_runtime_paths(
                frozen=True,
                executable=app_dir/"TVC Bot Worker.exe",
                bundle_dir=app_dir/"worker bundle",
            )
            _cfg,probe_path=tvc_probe._load_probe_config(gui_paths)
            self.assertEqual(gui_paths.config_file,external)
            self.assertEqual(worker_paths.config_file,external)
            self.assertEqual(probe_path,external)

    def test_gui_frozen_launch_uses_adjacent_worker_without_venv(self):
        with tempfile.TemporaryDirectory(prefix="tvc frozen gui ") as tmp:
            app_dir=Path(tmp)/"App With Spaces"
            app_dir.mkdir(parents=True)
            layout=resolve_runtime_paths(
                frozen=True,
                executable=app_dir/"TVC Bot Control.exe",
                bundle_dir=app_dir/"bundle",
            )
            worker=layout.worker_executable
            worker.write_bytes(b"mock worker")
            excel=app_dir/"Input Files"/"jobs one.xlsx"
            app=make_controller_app([excel])
            app.runtime_paths=layout
            app.bot_python=worker
            app.queue_running=True
            app.events=queue.Queue()
            app.current_file_index=-1
            app.current_file_var=Value("- / 1")
            app.current_job_var=Value("- / 0")
            app.current_job_ref=""
            app.stop_request_sent=False
            app.stop_event_seen=False
            app.stop_event_phase=""
            app.last_batch_success=None
            app.force_stop_used=False
            app.force_retry_available=False
            app.last_return_code=None
            app._refresh_queue_tree=mock.Mock()
            app._refresh_progress=mock.Mock()
            app._update_buttons=mock.Mock()
            process=mock.Mock()
            process.poll.return_value=None
            with (
                mock.patch.object(tvc_control.subprocess,"Popen",return_value=process) as popen,
                mock.patch.object(tvc_control.threading,"Thread") as thread,
            ):
                app._start_queue_item(0)
            command=popen.call_args.args[0]
            self.assertEqual(command[0],str(worker))
            self.assertNotIn(".venv"," ".join(command).lower())
            self.assertNotIn("bot.py"," ".join(command).lower())
            thread.return_value.start.assert_called_once()
            if app.stop_file is not None:
                app.stop_file.unlink(missing_ok=True)

    def test_missing_frozen_worker_blocks_runtime(self):
        with tempfile.TemporaryDirectory(prefix="tvc missing worker ") as tmp:
            app_dir=Path(tmp)
            layout=resolve_runtime_paths(
                frozen=True,
                executable=app_dir/"TVC Bot Control.exe",
                bundle_dir=app_dir,
            )
            with self.assertRaisesRegex(RuntimeError,"Worker executable"):
                tvc_control.validate_bot_runtime(runtime_paths=layout)

            app=make_controller_app([app_dir/"jobs.xlsx"])
            app.runtime_paths=layout
            app.root=mock.Mock()
            app.runtime_valid=False
            app.bot_python=None
            app.stop_request_sent=False
            for name in (
                "start_button","stop_button","browse_button","remove_button",
                "clear_button","up_button","down_button","runtime_button",
                "safety_button",
            ):
                setattr(app,name,mock.Mock())
            with mock.patch.object(tvc_control.messagebox,"showerror") as dialog:
                app._on_runtime_check_result(
                    False,None,f"ไม่พบ Worker executable: {layout.worker_executable}"
                )
            tvc_control.TVCControlApp._update_buttons(app)
            self.assertEqual(app.status_var.get(),"Error")
            self.assertEqual(
                app.start_button.configure.call_args.kwargs["state"],"disabled"
            )
            self.assertIn("Worker",dialog.call_args.args[0])


class TvcProbeTests(unittest.TestCase):
    @staticmethod
    def write_config(path,backend="win32"):
        path.parent.mkdir(parents=True,exist_ok=True)
        path.write_text(
            "[tvc]\n"
            "window_title_regex = ^เพิ่มใบงาน \\(JOB\\)$\n"
            f"backend = {backend}\n",
            encoding="utf-8",
        )

    def test_probe_source_config_uses_runtime_resolver(self):
        with tempfile.TemporaryDirectory(prefix="tvc probe source paths ") as tmp:
            root=Path(tmp)/"Project With Spaces"
            config_file=root/"config.ini"
            self.write_config(config_file,"win32")
            paths=resolve_runtime_paths(
                frozen=False,module_file=root/"src"/"runtime_paths.py"
            )
            cfg,resolved=tvc_probe._load_probe_config(paths)
            self.assertEqual(resolved,config_file)
            self.assertEqual(cfg.get("tvc","backend"),"win32")

    def test_probe_frozen_prefers_external_config_next_to_executable(self):
        with tempfile.TemporaryDirectory(prefix="tvc probe frozen paths ") as tmp:
            root=Path(tmp)/"Install Folder With Spaces"
            bundle=root/"bundle resources"
            external=root/"config.ini"
            bundled=bundle/"config.ini"
            self.write_config(bundled,"win32")
            self.write_config(external,"uia")
            paths=resolve_runtime_paths(
                frozen=True,
                executable=root/"TVC Bot Control.exe",
                bundle_dir=bundle,
            )
            cfg,resolved=tvc_probe._load_probe_config(paths)
            self.assertEqual(resolved,external)
            self.assertEqual(cfg.get("tvc","backend"),"uia")

    def test_gui_worker_and_probe_share_config_path(self):
        probe_cfg,path=tvc_probe._load_probe_config(tvc_control.RUNTIME_PATHS)
        self.assertTrue(probe_cfg.has_section("tvc"))
        self.assertEqual(path,tvc_control.RUNTIME_PATHS.config_file)
        self.assertEqual(path,bot.RUNTIME_PATHS.config_file)

    def test_probe_missing_config_is_explicit_error(self):
        with tempfile.TemporaryDirectory(prefix="tvc probe missing config ") as tmp:
            root=Path(tmp)/"Missing Config App"
            root.mkdir(parents=True)
            paths=resolve_runtime_paths(
                frozen=True,
                executable=root/"TVC Bot Control.exe",
                bundle_dir=root/"bundle",
            )
            with self.assertRaisesRegex(FileNotFoundError,"config.ini"):
                tvc_probe._load_probe_config(paths)

            result=tvc_control.perform_precheck(
                [],
                runtime_validator=lambda:Path("python.exe").resolve(),
                tvc_checker=lambda _runtime:tvc_probe.probe(paths),
                workbook_validator=lambda _path:True,
                access_checker=lambda _path:True,
                stats_reader=lambda _path:stats(wait=1),
                errors_reader=lambda _path:[],
            )
            self.assertFalse(result["tvc"]["ready"])
            self.assertIn("config.ini",result["tvc"]["message"])

    def test_probe_is_read_only_and_job_window_verifies_login(self):
        class Window:
            def window_text(self):
                return "เพิ่มใบงาน (JOB)"
            def children(self):
                raise AssertionError("JOB title match must not need a control walk")

        desktop=mock.Mock()
        desktop.windows.return_value=[Window()]
        with mock.patch.object(tvc_probe,"Desktop",return_value=desktop):
            result=tvc_probe.probe()
        self.assertEqual(result["status"],"READY")
        self.assertTrue(result["connected"])
        self.assertTrue(result["login_verified"])
        self.assertTrue(result["active_job_form"])
        self.assertEqual(desktop.windows.call_count,1)

    def test_probe_detects_reusable_job_form_by_controls_without_clicking(self):
        class Info:
            def __init__(self,automation_id="",process_id=42):
                self.automation_id=automation_id
                self.process_id=process_id

        class Control:
            def __init__(self,automation_id):
                self.element_info=Info(automation_id)
            def children(self):
                return []

        class MainWindow:
            element_info=Info()
            def window_text(self):
                return "T.V.C Client [Version 1.8.2]"
            def class_name(self):
                return "WindowsForms10.Window"

        class JobWindow:
            element_info=Info()
            def window_text(self):
                return "ข้อมูลรถ - งานปัจจุบัน"
            def class_name(self):
                return "WindowsForms10.Window"
            def children(self):
                return [Control("ButtonX3"),Control("ListView1"),Control("Tno")]

        win32_desktop=mock.Mock()
        win32_desktop.windows.return_value=[MainWindow()]
        uia_desktop=mock.Mock()
        uia_desktop.windows.return_value=[JobWindow()]

        def desktop_factory(*,backend):
            return uia_desktop if backend=="uia" else win32_desktop

        with mock.patch.object(tvc_probe,"Desktop",side_effect=desktop_factory):
            result=tvc_probe.probe()
        self.assertEqual(result["status"],"READY")
        self.assertTrue(result["login_verified"])
        self.assertIn("active_job_form_controls",result["login_signals"])
        self.assertTrue(result["active_job_form"])
        uia_desktop.windows.assert_called_once_with(process=42)

    def test_client_before_login_without_positive_signal_is_login_required(self):
        class Info:
            process_id=51
            automation_id=""

        class Window:
            element_info=Info()
            def __init__(self,title=""):
                self.title=title
            def window_text(self):
                return self.title
            def class_name(self):
                return "WindowsForms10.Window"
            def children(self):
                return []

        win32=mock.Mock()
        win32.windows.return_value=[Window("T.V.C Client [Version 1.8.2]")]
        uia=mock.Mock()
        uia.windows.return_value=[Window("T.V.C Client [Version 1.8.2]")]
        result=tvc_probe.probe(
            desktop_factory=lambda *,backend:uia if backend=="uia" else win32
        )
        self.assertEqual(result["status"],"LOGIN_REQUIRED")
        self.assertTrue(result["connected"])
        self.assertFalse(result["login_verified"])
        self.assertEqual(result["login_signals"],[])

    def test_logged_in_user_title_is_a_positive_signal(self):
        class Info:
            process_id=52

        class MainWindow:
            element_info=Info()
            def window_text(self):
                return "T.V.C Client [Version 1.8.2] เข้าสู่ระบบโดย ADMINPREMIUM []"
            def class_name(self):
                return "WindowsForms10.Window"

        win32=mock.Mock()
        win32.windows.return_value=[MainWindow()]
        uia=mock.Mock()
        uia.windows.side_effect=AssertionError(
            "verified logged-in title must not trigger UIA control enumeration"
        )
        result=tvc_probe.probe(
            desktop_factory=lambda *,backend:uia if backend=="uia" else win32
        )
        self.assertEqual(result["status"],"READY")
        self.assertTrue(result["login_verified"])
        self.assertIn("logged_in_user_title",result["login_signals"])
        uia.windows.assert_not_called()

    def test_post_login_job_menu_control_makes_neutral_title_ready(self):
        class Info:
            def __init__(self,process_id=53):
                self.process_id=process_id
                self.automation_id=""

        class Control:
            element_info=Info()
            def window_text(self):
                return "ใบงาน"
            def children(self):
                return []

        class Window:
            element_info=Info()
            def __init__(self,title=""):
                self.title=title
            def window_text(self):
                return self.title
            def class_name(self):
                return "WindowsForms10.Window"
            def children(self):
                return [Control()]

        win32=mock.Mock()
        win32.windows.return_value=[Window("T.V.C Client [Version 1.8.2]")]
        uia=mock.Mock()
        uia.windows.return_value=[Window("หน้าหลัก")]
        result=tvc_probe.probe(
            desktop_factory=lambda *,backend:uia if backend=="uia" else win32
        )
        self.assertEqual(result["status"],"READY")
        self.assertTrue(result["login_verified"])
        self.assertIn("post_login_control:ใบงาน",result["login_signals"])

    def test_login_word_alone_is_not_a_positive_signal(self):
        class Info:
            process_id=54
            automation_id=""

        class Window:
            element_info=Info()
            def window_text(self):
                return "T.V.C Client - Login"
            def class_name(self):
                return "WindowsForms10.Window"
            def children(self):
                return []

        win32=mock.Mock()
        win32.windows.return_value=[Window()]
        uia=mock.Mock()
        uia.windows.return_value=[Window()]
        result=tvc_probe.probe(
            desktop_factory=lambda *,backend:uia if backend=="uia" else win32
        )
        self.assertEqual(result["status"],"LOGIN_REQUIRED")
        self.assertFalse(result["login_verified"])

    def test_probe_not_found_is_fast_and_never_inspects_unrelated_controls(self):
        class UnrelatedWindow:
            def window_text(self):
                return "Visual Studio Code"
            def class_name(self):
                return "Chrome_WidgetWin_1"
            def children(self):
                raise AssertionError("unrelated controls must not be inspected")

        desktop=mock.Mock()
        desktop.windows.return_value=[UnrelatedWindow()]
        started=time.monotonic()
        with mock.patch.object(tvc_probe,"Desktop",return_value=desktop):
            result=tvc_probe.probe(timeout_seconds=0.5)
        elapsed=time.monotonic()-started
        self.assertEqual(result["status"],"NOT_FOUND")
        self.assertFalse(result["connected"])
        self.assertLess(elapsed,0.5)
        self.assertEqual(desktop.windows.call_count,1)

    def test_probe_internal_deadline_returns_timeout(self):
        class Clock:
            def __init__(self):
                self.value=0.0
            def __call__(self):
                self.value+=0.2
                return self.value

        desktop=mock.Mock()
        desktop.windows.return_value=[]
        result=tvc_probe.probe(
            timeout_seconds=0.1,
            desktop_factory=lambda **_kwargs:desktop,
            clock=Clock(),
        )
        self.assertEqual(result["status"],"TIMEOUT")
        self.assertTrue(result["timed_out"])


class CliCompatibilityTests(unittest.TestCase):
    def test_no_excel_argument_still_uses_config_without_connecting_tvc(self):
        with mock.patch.object(bot,"load_jobs",return_value=[]) as load_jobs:
            return_code=bot.main([])
        self.assertEqual(return_code,0)
        excel_path=load_jobs.call_args.args[0]
        self.assertEqual(
            excel_path,
            (Path(__file__).resolve().parents[1]/"data"/"TVC_Bot_JOB_Template_v5.xlsx").resolve(),
        )

    def test_explicit_excel_argument_remains_supported(self):
        parsed=bot.parse_args(["--excel","example.xlsx"])
        self.assertEqual(parsed.excel,"example.xlsx")


class AssetAndSummaryTests(unittest.TestCase):
    def test_version_source_of_truth(self):
        self.assertEqual(APP_NAME,"T.V.C JOB BOT")
        self.assertEqual(APP_VERSION,"0.8.0")

    def test_assets_existing_missing_and_invalid_never_crash(self):
        with tempfile.TemporaryDirectory(prefix="tvc_assets_") as tmp:
            existing=Path(tmp)/"image.png"
            existing.write_bytes(b"fixture")
            image=object()
            factory=mock.Mock(return_value=image)
            self.assertIs(
                tvc_control.load_photo_asset(existing,image_factory=factory),
                image,
            )
            self.assertIsNone(
                tvc_control.load_photo_asset(Path(tmp)/"missing.png",image_factory=factory)
            )
            invalid_factory=mock.Mock(side_effect=tk.TclError("invalid image"))
            self.assertIsNone(
                tvc_control.load_photo_asset(existing,image_factory=invalid_factory)
            )

    def test_summary_all_success_one_error_uncertain_and_stopped(self):
        done=ExcelQueueItem("done.xlsx",status="DONE",stats=stats(done=2))
        text,summary=tvc_control.format_queue_summary([done],"COMPLETE")
        self.assertIn("1/1 completed",text)
        self.assertEqual(summary["files_error"],0)

        error=ExcelQueueItem(
            "error.xlsx",
            status="ERROR",
            stats=stats(error=1),
            message="UNCERTAIN_TVC_SAVE",
            error_jobs=[{"job_ref":"TEST-014","bot_result":"UNCERTAIN_TVC_SAVE"}],
        )
        text,summary=tvc_control.format_queue_summary([done,error],"ERROR")
        self.assertEqual(summary["files_error"],1)
        self.assertIn("TEST-014",text)
        self.assertIn("UNCERTAIN_TVC_SAVE",text)

        stopped=ExcelQueueItem("stopped.xlsx",status="STOPPED",stats=stats(wait=3))
        text,summary=tvc_control.format_queue_summary([done,stopped],"STOPPED")
        self.assertIn("Queue Stopped",text)
        self.assertEqual(summary["WAIT"],3)

    def test_same_basename_in_two_directories_counts_as_two_error_files(self):
        first=ExcelQueueItem(Path("a")/"jobs.xlsx",status="ERROR",stats=stats(error=1))
        second=ExcelQueueItem(Path("b")/"jobs.xlsx",status="ERROR",stats=stats(error=1))
        summary=build_queue_summary([first,second])
        self.assertEqual(summary["files_error"],2)


if __name__=="__main__":
    unittest.main(verbosity=2)
